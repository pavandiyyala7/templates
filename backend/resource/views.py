from rest_framework import generics, filters, status
from rest_framework.views import APIView
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination
from django.core.cache import cache
from django.db.models import Q, F, Count, Sum
from django.db.models.signals import post_save, post_delete
from django.dispatch import receiver
from datetime import datetime, timedelta, date
from typing import Optional, Tuple, Any, List, Dict
import pytz
from dateutil import parser
from django.utils.timezone import make_aware, timezone, now
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.views.generic import View
from django.core.management import execute_from_command_line
from django.core.management import call_command
from django.db import transaction
from apscheduler.schedulers.background import BackgroundScheduler
from django_apscheduler.jobstores import DjangoJobStore
import time
import json
from django.http import StreamingHttpResponse
from io import BytesIO
from django.http import JsonResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side
from collections import defaultdict
from calendar import monthrange

from resource.models import Employee, Attendance, Logs, LastLogId,ManDaysAttendance, ManDaysMissedPunchAttendance, LastLogIdMandays, OvertimeRoundoffRules, ManualLogs, HolidayList
from resource.scheduler import get_scheduler
from . import serializers
from .services import generate_unique_ids, check_employee_id

from config.models import Company, Location

from resource import attendance5

class DefaultPagination(PageNumberPagination):
    """
    Default pagination class with page size set to 10.
    """
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class EmployeeIdGet(APIView):
    """
    API view for generating and retrieving employee IDs.
    """

    def get(self, request, *args, **kwargs):
        """
        Generate unique employee ID and device enroll ID.
        """
        employee_id, device_enroll_id = generate_unique_ids()
        return Response({"employee_id": employee_id, "device_enroll_id": device_enroll_id}, status=status.HTTP_200_OK)


# @receiver(post_save, sender=Employee)
# @receiver(post_delete, sender=Employee)
# def invalidate_employee_cache(sender, instance, **kwargs):
#     """
#     Signal handler to invalidate and reload the employee cache.
#     """
#     # Invalidate employee cache
#     cache.delete('employees')
#     cache.set('employees', Employee.objects.order_by('-id').all())

# @receiver(post_save, sender=Employee)
# def reload_employee_cache(sender, instance, created, **kwargs):
#     """
#     Signal handler to reload the employee cache when a new record is added or updated.
#     """
#     # Reload employee cache only if a new record is created or an existing record is updated
#     if created or not instance._state.adding:
#         cache.set('employees', Employee.objects.order_by('-id').all())


# # Pre-load data into cache
# cache.get_or_set('employees', Employee.objects.order_by('-id').all(), timeout=3600)


class EmployeeListCreate(generics.ListCreateAPIView):
    """
    API view for listing and creating employees.
    """
    queryset = Employee.objects.all()
    serializer_class = serializers.EmployeeSerializer
    pagination_class = DefaultPagination
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    filter_fields = '__all__'
    search_fields = ['employee_id', 'employee_name', 'job_status']
    ordering_fields = '__all__'
    ordering = ['-id']

    # def get_queryset(self):
    #     """
    #     Get the cached queryset for listing employees with optional search query.
    #     """
    #     queryset = cache.get('employees')
    #     if queryset is None:
    #         print("Fetching queryset from the database.")
    #         queryset = Employee.objects.order_by('-id').all()
    #         cache.set('employees', queryset)
    #     else:
    #         print("Fetching queryset from the cache.")
    #     return queryset

    def search_queryset(self, queryset, search_query):
        """
        Filter the queryset based on the search query.
        """
        return queryset.filter(
            Q(employee_id__icontains=search_query) |
            Q(employee_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            # Q(phone_no__icontains=search_query) |
            # Q(designation_name__icontains=search_query) |
            # Q(department_name__icontains=search_query) |
            # Q(location_name__icontains=search_query) |
            Q(job_status__icontains=search_query)
        )

    def get(self, request, *args, **kwargs):
        """
        Get the list of employees with optional search query.
        """
        search_query = self.request.GET.get('search')
        queryset = self.get_queryset()

        if search_query:
            queryset = self.search_queryset(queryset, search_query)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class EmployeeRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    """
    API view for retrieving, updating, and deleting an employee.
    """
    queryset = Employee.objects.all()
    serializer_class = serializers.EmployeeSerializer
    lookup_url_kwarg = "id"

    def perform_update(self, serializer):
        """
        Update the employee instance, converting shift 0 to None.
        """
        data = serializer.validated_data
        # If shift is 0, convert it to None
        if data.get('shift') == 0:
            data['shift'] = None
        serializer.save()


class AttendanceListCreate(generics.ListCreateAPIView):
    """
    API view for listing and creating attendance records.
    """
    # queryset = Attendance.objects.all()
    serializer_class = serializers.AttendanceSerializer
    pagination_class = DefaultPagination
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    filter_fields = '__all__'
    search_fields = ['employee_id', 'employee_name', 'device_enroll_id', 'logdate', 'shift', 
                    '=shift_status', 'first_logtime', 'last_logtime', 'total_time', 'late_entry', 
                    'early_exit', 'overtime', 'company_name', 'location_name']
    ordering_fields = '__all__'
    ordering = ['-logdate']

    def get_queryset(self):
        """
        Get the queryset for listing attendance records with optional search query.
        """
        search_query = self.request.GET.get('search')
        date_query = self.request.GET.get('date')

        # Get date range parameters
        date_from = self.request.GET.get('date_from') 
        date_to = self.request.GET.get('date_to') 

        shift_status = self.request.GET.get('shift_status')
        late_entry_not_null = self.request.GET.get('late_entry')
        early_exit_not_null = self.request.GET.get('early_exit')
        overtime_not_null = self.request.GET.get('overtime')
        missed_punch = self.request.GET.get('missed_punch')
        employee_id = self.request.GET.get('employeeid')  
        employee_ids = self.request.GET.get('employee_ids')
        company_names = self.request.GET.get('company_name')
        location_names = self.request.GET.get('location_name') 
        department_names = self.request.GET.get('department_name')
        designation_names = self.request.GET.get('designation_name')
        insufficient_duty_hours = self.request.GET.get('insufficient_duty_hours')

        queryset = Attendance.objects.order_by('-logdate').all()

        if search_query:
            queryset = queryset.filter(
                Q(employeeid__employee_name__icontains=search_query) |
                Q(employeeid__device_enroll_id__icontains=search_query) |
                Q(logdate__icontains=search_query) |
                Q(shift_status__iexact=search_query) |
                Q(first_logtime__icontains=search_query) |
                Q(last_logtime__icontains=search_query) |
                Q(total_time__icontains=search_query) |
                Q(late_entry__icontains=search_query) |
                Q(employeeid__company__name__icontains=search_query) |
                Q(employeeid__location__name__icontains=search_query) |
                Q(employeeid__department__name__icontains=search_query) |
                Q(employeeid__designation__name__icontains=search_query) 
            )

        if date_query:
            try:
                # Convert date_query to a datetime object
                date_obj = datetime.strptime(date_query, '%m-%d-%Y').date()
                queryset = queryset.filter(logdate=date_obj)
            except ValueError:
                # Handle invalid date format
                pass

        # Filter by date range if both date_from and date_to are provided
        if date_from and date_to:
            try:
                date_from_obj = datetime.strptime(date_from, '%m-%d-%Y').date()
                date_to_obj = datetime.strptime(date_to, '%m-%d-%Y').date()
                queryset = queryset.filter(logdate__range=[date_from_obj, date_to_obj]).order_by('logdate')
            except ValueError:
                pass 

        if shift_status:
            queryset = queryset.filter(shift_status=shift_status)

        # Filter by non-null late_entry values if late_entry_not_null parameter is provided
        if late_entry_not_null == 'true':
            queryset = queryset.exclude(late_entry__isnull=True).exclude(late_entry='00:00:00')
        if early_exit_not_null == 'true':
            queryset = queryset.exclude(early_exit__isnull=True).exclude(early_exit='00:00:00')
        if overtime_not_null == 'true':
            queryset = queryset.exclude(overtime__isnull=True).exclude(overtime='00:00:00')
        if missed_punch == 'true':
            queryset = queryset.filter(shift_status='MP')

        if insufficient_duty_hours == 'true':
            queryset = queryset.filter(total_time__lt='08:00:00')

        if employee_id:
            queryset = queryset.filter(employeeid=employee_id)

        if employee_ids:
            employee_ids_list = [id.strip() for id in employee_ids.split(',')]
            queryset = queryset.filter(employeeid__employee_id__in=employee_ids_list)
            print(employee_ids_list)

        if company_names:
            company_names_list = [name.strip() for name in company_names.split(',')]
            queryset = queryset.filter(employeeid__company__name__in=company_names_list)
        
        if location_names:
            location_names_list = [name.strip() for name in location_names.split(',')]
            queryset = queryset.filter(employeeid__location__name__in=location_names_list)

        if department_names:
            department_names_list = [name.strip() for name in department_names.split(',')]
            queryset = queryset.filter(employeeid__department__name__in=department_names_list)

        if designation_names:
            designation_names_list = [name.strip() for name in designation_names.split(',')]
            queryset = queryset.filter(employeeid__designation__name__in=designation_names_list)

        # Ensure consistent ordering by a specific field, for example, by 'id'
        # queryset = queryset.order_by('-logdate')

        return queryset

    def get(self, request, *args, **kwargs):
        """
        Get the list of attendance records with optional search query.
        """
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

# @receiver([post_save, post_delete], sender=Attendance)
# def invalidate_and_reload_attendance_cache(sender, instance, **kwargs):
#     """
#     Signal handler to invalidate and reload the employee cache.
#     """
#     cache.delete('attendance')
#     cache.set('attendance', Attendance.objects.order_by('-logdate').all())

# # Pre-load data into cache
# cache.get_or_set('attendance', Attendance.objects.order_by('-logdate').all(), timeout=3600)


class ExportAttendanceExcelView(View):
    HEADERS = (
        "Employee ID", "Device Enroll ID", "Employee Name", "Company", "Location", 
        "Job Type", "Department", "Employee Type", "Desination", "Log Date", 
        "Shift", "Shift Status", "In Time", "Out Time", "Total Hours", 
        "Late Entry", "Early Exit", "OT Hours"
    )

    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True)),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True)),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True)),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True)),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True)),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True)),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True)),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True)),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True)),
    }


    def get_filtered_queryset(self, request):
        """Get optimized and filtered queryset"""
        # Use select_related for all foreign key relationships
        # queryset = Attendance.objects.select_related(
        #     'employeeid',
        #     'employeeid__company',
        #     'employeeid__location',
        #     'employeeid__department',
        #     'employeeid__designation',
        #     'employeeid__shift'
        # ).order_by('-logdate')
        queryset = Attendance.objects.select_related(
            'employeeid',
            'employeeid__company',
            'employeeid__location',
            'employeeid__department',
            'employeeid__designation',
            'employeeid__shift'
        ).values(
            'employeeid__employee_id',
            'employeeid__device_enroll_id',
            'employeeid__employee_name',
            'employeeid__company__name',
            'employeeid__location__name',
            'employeeid__job_type',
            'employeeid__department__name',
            'employeeid__category',
            'employeeid__designation__name',
            'logdate',
            'shift',
            'shift_status',
            'first_logtime',
            'last_logtime',
            'total_time',
            'late_entry',
            'early_exit',
            'overtime'
        ).order_by('-logdate')

        # Date range filter
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        if date_from and date_to:
            try:
                date_from_obj = datetime.strptime(date_from, '%m-%d-%Y').date()
                date_to_obj = datetime.strptime(date_to, '%m-%d-%Y').date()
                queryset = queryset.filter(logdate__range=[date_from_obj, date_to_obj]).order_by('employeeid', 'logdate')
            except ValueError:
                pass
        
        employee_ids = request.GET.get('employee_ids')
        if employee_ids:
            employee_ids_list = [emp_id.strip() for emp_id in employee_ids.split(',') if emp_id.strip()]
            if employee_ids_list:
                queryset = queryset.filter(employeeid__employee_id__in=employee_ids_list)

        # Direct filters
        filters = {}
        if request.GET.get('employee_id'):
            filters['employeeid__employee_id__iexact'] = request.GET.get('employee_id')
        if request.GET.get('employee_name'):
            filters['employeeid__employee_name__icontains'] = request.GET.get('employee_name')
        if request.GET.get('shift_status'):
            filters['shift_status'] = request.GET.get('shift_status')

        # Apply direct filters
        if filters:
            queryset = queryset.filter(**filters)

        # Handle comma-separated values
        for param, field in {
            'employee_id': 'employeeid__employee_id__in',
            'company_name': 'employeeid__company__name__in',
            'location_name': 'employeeid__location__name__in',
            'department_name': 'employeeid__department__name__in',
            'designation_name': 'employeeid__designation__name__in'
        }.items():
            values = request.GET.get(param)
            if values:
                queryset = queryset.filter(**{field: [name.strip() for name in values.split(',')]})

        # Special filters
        if request.GET.get('late_entry') == 'true':
            queryset = queryset.exclude(late_entry__isnull=True).exclude(late_entry='00:00:00')
        if request.GET.get('early_exit') == 'true':
            queryset = queryset.exclude(early_exit__isnull=True).exclude(early_exit='00:00:00')
        if request.GET.get('overtime') == 'true':
            queryset = queryset.exclude(overtime__isnull=True).exclude(overtime='00:00:00')
        if request.GET.get('missed_punch') == 'true':
            queryset = queryset.exclude(first_logtime__isnull=True).exclude(last_logtime__isnull=False)
        if request.GET.get('insufficient_duty_hours') == 'true':
            queryset = queryset.filter(total_time__lt='08:00:00')

        return queryset
    
    def setup_worksheet(self, wb):
        """Initialize worksheet with headers and styling"""
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Write headers
        for col_num, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(size=14, bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            ws.column_dimensions[cell.column_letter].width = len(header) + 7

        ws.freeze_panes = 'A2'
        return ws

    def get_record_data(self, record):
        """Extract record data into tuple for faster access"""
        return (
            record['employeeid__employee_id'],  # Access flattened key
            record['employeeid__device_enroll_id'],
            record['employeeid__employee_name'],
            record['employeeid__company__name'],
            record['employeeid__location__name'],
            record['employeeid__job_type'],
            record['employeeid__department__name'],
            record['employeeid__category'],
            record['employeeid__designation__name'],
            record['logdate'],
            record['shift'],
            record['shift_status'],
            record['first_logtime'],
            record['last_logtime'],
            record['total_time'],
            record['late_entry'],
            record['early_exit'],
            record['overtime']
        )

    def write_records(self, ws, records):
        """Write all records to the worksheet in batches"""
        for record_data in records:
            ws.append(record_data)

    def apply_styles(self, ws, num_records):
        """Apply styles to the worksheet after writing data"""
        # Define a thin border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply center alignment and thin border to all cells
        for row in ws.iter_rows(min_row=1, max_row=num_records + 1, min_col=1, max_col=len(self.HEADERS)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')  # Center all cells
                cell.border = thin_border  # Apply thin border

        # Apply shift status styles to the "Shift Status" column (column 12)
        for row in ws.iter_rows(min_row=2, max_row=num_records + 1, min_col=12, max_col=12):
            for cell in row:
                status_style = self.SHIFT_STATUS_STYLES.get(cell.value)
                if status_style:
                    cell.style = status_style
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')

    # def write_record(self, ws, row_num, record_data):
    #     """Write a single record to worksheet with styling"""
    #     for col_num, value in enumerate(record_data, 1):
    #         cell = ws.cell(row=row_num, column=col_num, value=value)
    #         cell.alignment = Alignment(horizontal='center')
            
    #         # Apply shift status styling (column 12)
    #         if col_num == 12:  # Shift Status column
    #             status_style = self.SHIFT_STATUS_STYLES.get(value)
    #             if status_style:
    #                 cell.style = status_style
    #             cell.alignment = Alignment(horizontal='center')

    def get(self, request, *args, **kwargs):
        # Get filtered queryset with all related data
        queryset = self.get_filtered_queryset(request)
        
        # Convert queryset to tuple of tuples for better performance
        records = tuple(
            self.get_record_data(record) for record in queryset
        )

        # Create workbook and setup worksheet
        wb = openpyxl.Workbook()
        ws = self.setup_worksheet(wb)

        # Write all records
        self.write_records(ws, records)

        # Apply styles
        self.apply_styles(ws, len(records))

        # Create response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=Attendance_Report.xlsx"
        wb.save(response)

        return response
        
class AttendanceMetricsAPIView(generics.ListAPIView):
    def get_queryset(self):
        # Get the latest log date
        # latest_logdate = datetime.strptime('2023-08-03', '%Y-%m-%d').date()
        latest_logdate = Attendance.objects.latest('logdate').logdate
        # print(latest_logdate)
        # Filter the queryset to include only records with the latest log date
        queryset = Attendance.objects.filter(logdate=latest_logdate)
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        
        # Counting logic
        present_count = queryset.filter(shift_status='P').count()
        absent_count = queryset.filter(shift_status='A').count()
        late_entry_count = queryset.exclude(late_entry__isnull=True).count()
        early_exit_count = queryset.exclude(early_exit__isnull=True).count()
        overtime_count = queryset.exclude(overtime__isnull=True).count()

        latest_logdate = queryset.latest('logdate').logdate

        # counting present count since last hour(current hour - 1 hour)
        present_count_last_hour = queryset.filter(shift_status='P', first_logtime__gte=now()-timedelta(hours=1)).count()

        # Calculate the absent count for the previous week
        previous_week_absent_count = queryset.filter(shift_status='A', logdate__range=[latest_logdate - timedelta(days=7), latest_logdate - timedelta(days=1)]).count()

        # Calculate the absent count for the current week
        current_week_absent_count = queryset.filter(shift_status='A', logdate__range=[latest_logdate - timedelta(days=1), latest_logdate]).count()

        # Calculate the percentage increase in absentees compared to last week
        if previous_week_absent_count != 0:
            absent_percentage_increase = ((current_week_absent_count - previous_week_absent_count) / previous_week_absent_count) * 100
        else:
            absent_percentage_increase = 0

        # Counting employees who arrive late frequently
        late_arrival_threshold = 5  # Define the threshold for frequent late arrivals
        frequent_late_arrivals = queryset.filter(late_entry__isnull=False).annotate(late_count=Count('late_entry')).filter(late_count__gte=late_arrival_threshold).count()

        # Counting employees who arrived late today
        today = date.today()
        late_today = queryset.filter(late_entry__isnull=False, logdate=today).count()
        
        late_arrival_threshold = 5  # Define the threshold for frequent late arrivals
        frequent_late_arrivals = queryset.filter(late_entry__isnull=False).annotate(late_count=Count('late_entry')).filter(late_count__gte=late_arrival_threshold).count()

        current_hour = datetime.now().hour
        # Counting live headcount 
        if 5 <= current_hour <= 17:
            live_headcount = queryset.filter(first_logtime=F('last_logtime')).count()
        else:
            live_headcount = 0    
        # counting total checkin whose first logtime is not null
        total_checkin = queryset.exclude(first_logtime__isnull=True).count()
        # counting total checkout whose last logtime is not null and not equal to first logtime
        total_checkout = queryset.exclude(last_logtime__isnull=True).exclude(last_logtime=F('first_logtime')).count()
        
        # Constructing response data
        data = {
            'present_count': present_count,
            'absent_count': absent_count,
            'late_entry_count': late_entry_count,
            'early_exit_count': early_exit_count,
            'overtime_count': overtime_count,
            'live_headcount': live_headcount,
            'total_checkin': total_checkin,
            'total_checkout': total_checkout,
            'present_count_last_hour': present_count_last_hour,
            'absent_percentage_increase': absent_percentage_increase,
            'frequent_late_arrivals': frequent_late_arrivals,
        }
        
        return Response(data)
    
class AttendanceMonthlyMetricsAPIView(APIView):
    
    def get(self, request, format=None):
        current_date = datetime.now()
        start_date = current_date.replace(day=1)
        
        attendances = Attendance.objects.filter(logdate__range=[start_date, current_date])
        
        # Dictionary to keep track of metrics per day
        metrics_dict = {}

        for attendance in attendances:
            logdate_str = attendance.logdate.strftime('%Y-%m-%d')
            if logdate_str not in metrics_dict:
                metrics_dict[logdate_str] = {
                    'date': logdate_str,
                    'present': 0,
                    'absent': 0,
                    'late_entry': 0,
                    'early_exit': 0,
                    'overtime': 0
                }

            if attendance.shift_status == 'A':
                metrics_dict[logdate_str]['absent'] += 1
            else:
                metrics_dict[logdate_str]['present'] += 1

            if attendance.late_entry and attendance.late_entry != timedelta(0):
                metrics_dict[logdate_str]['late_entry'] += 1
            
            if attendance.early_exit and attendance.early_exit != timedelta(0):
                metrics_dict[logdate_str]['early_exit'] += 1
            
            if attendance.overtime and attendance.overtime != timedelta(0):
                metrics_dict[logdate_str]['overtime'] += 1

        # Convert metrics_dict to a list of daily metrics and sort by date
        daily_metrics = sorted(metrics_dict.values(), key=lambda x: x['date'])

        response_data = {
            'daily_metrics': daily_metrics
        }
        
        return Response(response_data, status=status.HTTP_200_OK)
    
class LogsListCreate(generics.ListCreateAPIView):
    """
    API view for listing and creating logs with improved filtering.
    """
    serializer_class = serializers.LogsSerializer
    pagination_class = DefaultPagination
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    filter_fields = '__all__'
    search_fields = ['employeeid', 'shortname']
    ordering_fields = ['log_datetime', 'employeeid', 'direction', 'shortname']
    ordering_fields = '__all__'
    ordering = ['-log_datetime']

    def get_queryset(self):
        """
        Get the queryset for listing logs with all filters.
        """
        queryset = Logs.objects.all()

        # Get query parameters
        search_query = self.request.GET.get('search')
        date_query = self.request.GET.get('date')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        employee_id = self.request.GET.get('employeeid')
        employee_ids = self.request.GET.get('employee_ids')
        direction = self.request.GET.get('direction')
        
        # Apply search filter (case-insensitive)
        if search_query:
            queryset = queryset.filter(
                Q(employeeid__icontains=search_query) |
                Q(shortname__icontains=search_query)
            )

        # Filter by specific date
        if date_query:
            try:
                date_obj = datetime.strptime(date_query, '%Y-%m-%d').date()
                queryset = queryset.filter(log_datetime__date=date_obj)
            except ValueError:
                pass

        # Filter by date range
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(log_datetime__date__gte=date_from_obj)
            except ValueError:
                pass

        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(log_datetime__date__lte=date_to_obj)
            except ValueError:
                pass

        # Filter by single employee ID
        if employee_id:
            queryset = queryset.filter(employeeid=employee_id)

        # Filter by multiple employee IDs
        if employee_ids:
            try:
                employee_ids_list = [id.strip() for id in employee_ids.split(',')]
                queryset = queryset.filter(employeeid__in=employee_ids_list)
            except Exception:
                pass

        # Filter by direction
        if direction:
            queryset = queryset.filter(direction=direction)

        return queryset.order_by('-log_datetime')

    def get(self, request, *args, **kwargs):
        """
        Get the list of Logs records with filters.
        """
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
            
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)        

class LogsRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    """
    API view for retrieving, updating, and deleting a log.
    """
    queryset = Logs.objects.all()
    serializer_class = serializers.LogsSerializer
    lookup_url_kwarg = "id"

class EmployeeDropdownList(generics.ListAPIView):
    """
    API view for listing employees in a dropdown.
    """
    queryset = Employee.objects.all().order_by('employee_id')
    pagination_class = None
    serializer_class = serializers.EmployeeDropdownSerializer

class ExportEmployeeAttendanceExcelView(View):
    def get(self, request, *args, **kwargs):
        employee_id = request.GET.get('employee_id')
        month = request.GET.get('month')
        year = request.GET.get('year')

        queryset = Attendance.objects.order_by('-logdate').all()

        if employee_id:
            queryset = queryset.filter(Q(employeeid__employee_id__iexact=employee_id))
        if month:
            queryset = queryset.filter(logdate__month=month)
        if year:
            queryset = queryset.filter(logdate__year=year)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Employee {employee_id} Attendance Report"

        # Create the row with the required employee details
        if employee_id and month and year:
            employee = queryset.first().employeeid
            first_day_of_month = date(int(year), int(month), 1)
            next_month = first_day_of_month.replace(day=28) + timedelta(days=4)
            last_day_of_month = next_month - timedelta(days=next_month.day)
            num_days = last_day_of_month.day

            # ws.append([f"Monthly In and Out Register for the Period of: {first_day_of_month} to {last_day_of_month}"])
            # Merge cells for the header
            header_text = f"Monthly In and Out Register for the Period of: {first_day_of_month} to {last_day_of_month}"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
            header_cell = ws.cell(row=1, column=1, value=header_text)
            ws.append([f"EMP ID: ", f"{employee.employee_id}", "CALENDAR DAYS:", num_days, "Days"] + [str(day) for day in range(1, num_days + 1)])

            # Fetch the shift data for each day
            shifts = []
            shift_status = []
            first_logtime = []
            last_logtime = []
            total_hours = []
            late_entries = []
            early_exits = []
            overtime = []

            for day in range(1, num_days + 1):
                log_date = date(int(year), int(month), day)
                attendance_record = queryset.filter(logdate=log_date).first()
                if attendance_record:
                    shifts.append(attendance_record.shift if attendance_record.shift else "")
                    if attendance_record.shift_status == 'P':
                        shift_status.append(attendance_record.shift_status)
                    elif attendance_record.shift_status == 'A':
                        if log_date.weekday() >= 6:
                            shift_status.append("WO")
                        else:
                            shift_status.append("A")
                    first_logtime.append(attendance_record.first_logtime if attendance_record.first_logtime else "")
                    last_logtime.append(attendance_record.last_logtime if attendance_record.last_logtime else "")
                    total_hours.append(attendance_record.total_time if attendance_record.total_time else "")
                    late_entries.append(attendance_record.late_entry if attendance_record.late_entry else "")
                    early_exits.append(attendance_record.early_exit if attendance_record.early_exit else "")
                    overtime.append(attendance_record.overtime if attendance_record.overtime else "")
                else:
                    shifts.append("")
                    shift_status.append("")
                    first_logtime.append("")
                    last_logtime.append("")
                    total_hours.append("")
                    late_entries.append("")
                    early_exits.append("")
                    overtime.append("")

            total_present_days = shift_status.count('P') + shift_status.count('HD')  # Count total "P" and "HD" days

            ws.append([f"EMP Name: ", f"{employee.employee_name}", "PAID DAYS:", "", "Shift"] + shifts)
            ws.append([f"COMPANY: ", f"{employee.company.name if employee.company else ''}", "PRESENT DAYS:", f"{total_present_days}", "Status"] + shift_status)
            for idx, status in enumerate(shift_status, start=6):  # Adjust the starting index based on the row where 'shift_status' starts
                cell = ws.cell(row=4, column=idx)  # Adjust the row based on where 'shift_status' is located
                if status == 'P':
                    cell.style = 'Good'
                elif status == 'WO': 
                    cell.style = 'Neutral'
                else:
                    cell.style = 'Bad'

                cell.alignment = Alignment(horizontal='center')

            total_WO_days = shift_status.count('WO') + shift_status.count('WW')  # Count total "WO" days
            total_late_entry_time  = sum([t for t in late_entries if t], timedelta())
            total_early_exit_time = sum([t for t in early_exits if t], timedelta())
            total_OT_time = sum([t for t in overtime if t], timedelta())
            total_working_time = sum([t for t in total_hours if t], timedelta())

            total_GS_days = shifts.count('GS') if shifts else 0
            total_OS_days = shifts.count('OS') if shifts else 0
            total_FS_days = shifts.count('FS') if shifts else 0
            total_SS_days = shifts.count('SS') if shifts else 0
            total_NS_days = shifts.count('NS') if shifts else 0

            ws.append(["LOCATION: ", f"{employee.location.name if employee.location else ''}", "WO:", f"{total_WO_days}", "Duty-In"] + first_logtime)
            ws.append(["DEPARTMENT: ", f"{employee.department.name if employee.department else ''}", "WW:", "0", "Duty-Out"] + last_logtime)
            ws.append(["DESIGNATION: ", f"{employee.designation.name if employee.designation else ''}", "FS:", "0", "Duty Hours"] + total_hours)
            ws.append(["EMP TYPE: ", f"{employee.job_type if employee.job_type else ''}", "PH:", "0", "Lunch-Out"] + [" "] * num_days)
            ws.append(["REPORTING MNG: ", f"{employee.reporting_manager.employee_name [employee.reporting_manager.employee_id] if employee.reporting_manager else ''}", "CO:", "0", "Lunch-In"] + [" "] * num_days)
            ws.append(["CL:", "0", "CW:", "0", "Lunch Hours"] + [" "] * num_days)
            ws.append(["EL:", "0", "LATE ENTRY HOURS:", f"{total_late_entry_time}", "OT-In"])
            ws.append(["SL:", "0", "EARLY EXIT HOURS:", f"{total_early_exit_time}", "OT-Out"])
            ws.append(["MEDICAL:", "0", "OT HOURS:", f"{total_OT_time}", "OT Hours"] + overtime)
            ws.append(["ON DUTY:", "0", "LUNCH HOURS:", "", "Late Entry"] + late_entries)
            ws.append([f"GS:{total_GS_days} OS:{total_OS_days} FS:{total_FS_days} SS:{total_SS_days} NS:{total_NS_days}", "", "WORKING HOURS:", f"{total_working_time}", "Early Exit"] + early_exits)
        
        # Auto-adjust column widths with a minimum width
        min_width = 10  # Define the minimum width for the columns
        for col_idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)  # Get the column letter using index

            # Iterate through all cells in the column, skipping the first row
            for cell in column[1:]:  # Skip first row
                if cell.row == 1 and cell.column < 6:  # Skip merged header cells if needed
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Adjust column width, ensuring it meets the minimum width
            adjusted_width = max(max_length + 2, min_width)  # Add padding and ensure it meets minimum width
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Convert the month and year to a datetime object
        month_name = datetime(year=int(year), month=int(month), day=1).strftime('%B')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_{employee.employee_id}_Attendance_{month_name}_{year}.xlsx'
        wb.save(response)
        return response
    

class ExportAllEmployeeAttendanceExcelView(View):
    """
    View to generate and export detailed attendance Excel report for all employees.
    Includes comprehensive attendance data with styled formatting.
    """

    @property
    def SHIFT_STATUS_STYLES(self):
        """Define and register named styles for shift statuses."""
        if not hasattr(self, '_shift_status_styles'):
            self._shift_status_styles = {
                'P': {
                    'fill': PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
                    'font': Font(color="256029", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'WW': {
                    'fill': PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
                    'font': Font(color="256029", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'A': {
                    'fill': PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
                    'font': Font(color="C63737", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'HD': {
                    'fill': PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"),
                    'font': Font(color="8A5340", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'WO': {
                    'fill': PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"),
                    'font': Font(color="8A5340", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'PH': {
                    'fill': PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"),
                    'font': Font(color="8A5340", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'FH': {
                    'fill': PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"),
                    'font': Font(color="8A5340", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'MP': {
                    'fill': PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),
                    'font': Font(color="D84315", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'IH': {
                    'fill': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
                    'font': Font(color="856404", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                }
            }
        return self._shift_status_styles

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """Handle GET request to generate the Excel file."""
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse(
                {'error': 'Month and year are required parameters.'}, 
                status=400
            )

        # Initialize date parameters
        first_day = date(int(year), int(month), 1)
        last_day = self.get_last_day_of_month(first_day)
        num_days = last_day.day

        # Fetch all required data at once
        attendance_data, employee_data = self.fetch_data(month, year)

        # Generate Excel
        wb = self.generate_excel(
            attendance_data, 
            employee_data, 
            first_day, 
            last_day, 
            num_days
        )

        # Prepare response
        month_name = first_day.strftime('%B')
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename=All_Employee_Attendance_{month_name}_{year}.xlsx'
        )
        wb.save(response)
        return response

    def fetch_data(self, month: str, year: str) -> Tuple[Dict, Dict]:
        """
        Fetch all required attendance and employee data efficiently.
        """
        attendance_records = (
            Attendance.objects.filter(
                logdate__year=year, 
                logdate__month=month
            )
            .select_related('employeeid')
            .order_by('employeeid', 'logdate')
        )
        
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        employees = (
            Employee.objects.all()
            .select_related(
                'department', 
                'designation', 
                'location', 
                'company', 
                'reporting_manager'
            )
        )
        
        employee_data = {
            emp.id: emp for emp in employees
        }

        return attendance_data, employee_data

    def generate_excel(
        self, 
        attendance_data: Dict, 
        employee_data: Dict, 
        first_day: date, 
        last_day: date, 
        num_days: int
    ) -> Workbook:
        """Generate Excel workbook with attendance data."""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{first_day.month}_{first_day.year}"

        row_num = 1
        for emp_id, employee in employee_data.items():
            if emp_id in attendance_data:
                self.add_employee_attendance(
                    ws,
                    employee,
                    attendance_data[emp_id],
                    first_day,
                    last_day,
                    num_days,
                    row_num
                )
                row_num += 16

        self.auto_adjust_column_width(ws)
        return wb

    def add_employee_attendance(
        self, 
        ws, 
        employee, 
        attendance_records: List, 
        first_day: date, 
        last_day: date, 
        num_days: int, 
        start_row: int
    ):
        """Add employee attendance data to worksheet with proper formatting."""
        attendance_data = self.process_attendance_records(
            attendance_records, 
            num_days, 
            first_day
        )
        
        # Add employee details rows
        self.add_employee_details_rows(
            ws, 
            employee, 
            attendance_data, 
            num_days, 
            start_row
        )

        # Apply styles
        self.apply_cell_styles(ws, start_row, num_days)

    def process_attendance_records(
        self, 
        records: List, 
        num_days: int, 
        first_day: date
    ) -> Dict:
        """
        Process attendance records into required format.
        
        Args:
            records: List of attendance records
            num_days: Number of days in the month
            first_day: First day of the month
            
        Returns:
            Dict containing processed attendance data
        """
        # Initialize data structures
        attendance_map = {
            record.logdate.day: record 
            for record in records
        }
        
        # Initialize attendance data
        data = {
            'shifts': [],
            'shift_status': [],
            'first_logtime': [],
            'last_logtime': [],
            'total_hours': [],
            'late_entries': [],
            'early_exits': [],
            'overtime': []
        }
        
        # Process each day
        for day in range(1, num_days + 1):
            record = attendance_map.get(day)
            
            if record:
                data['shifts'].append(record.shift or "")
                data['shift_status'].append(record.shift_status or "")
                data['first_logtime'].append(record.first_logtime or "")
                data['last_logtime'].append(record.last_logtime or "")
                data['total_hours'].append(record.total_time or "")
                data['late_entries'].append(record.late_entry or "")
                data['early_exits'].append(record.early_exit or "")
                data['overtime'].append(record.overtime or "")
            else:
                # Add empty values for missing records
                for key in data:
                    data[key].append("")
        
        # Calculate totals
        data['total_present_days'] = (
            data['shift_status'].count('P') + 
            data['shift_status'].count('HD')
        )
        data['total_WO_days'] = data['shift_status'].count('WO') + data['shift_status'].count('WW')
        data['total_WW_days'] = data['shift_status'].count('WW')
        
        # Calculate time totals
        data['total_late_entry_time'] = sum(
            [t for t in data['late_entries'] if t], 
            timedelta()
        )
        data['total_early_exit_time'] = sum(
            [t for t in data['early_exits'] if t], 
            timedelta()
        )
        data['total_OT_time'] = sum(
            [t for t in data['overtime'] if t], 
            timedelta()
        )
        data['total_working_time'] = sum(
            [t for t in data['total_hours'] if t], 
            timedelta()
        )
        
        return data

    def add_employee_details_rows(
        self, 
        ws, 
        employee, 
        attendance_data: Dict, 
        num_days: int, 
        start_row: int
    ):
        """
        Add employee details and attendance rows to worksheet.
        
        Args:
            ws: Worksheet object
            employee: Employee object
            attendance_data: Processed attendance data
            num_days: Number of days in month
            start_row: Starting row number
        """
        days = [str(day) for day in range(1, num_days + 1)]
        
        # Row 1: Employee Header - Stored separately for width adjustment exclusion
        header_text = (
            f"EMP ID: {employee.employee_id} | EMP Name: {employee.employee_name} | "
            f"Department: {employee.department.name if employee.department else ''} | "
            f"Designation: {employee.designation.name if employee.designation else ''} | "
            f"Type: {employee.job_type if employee.job_type else ''} | Location: {employee.location.name if employee.location else ''} | "
            f"Company: {employee.company.name if employee.company else ''}"
        )
        
        # Store header information for width adjustment exclusion
        if not hasattr(ws, '_header_rows'):
            ws._header_rows = set()
        ws._header_rows.add(start_row)
        
        # Merge and set header
        ws.merge_cells(
            start_row=start_row, 
            start_column=1, 
            end_row=start_row, 
            end_column=num_days + 5
        )
        header_cell = ws.cell(row=start_row, column=1, value=header_text)
        header_cell.font = Font(bold=True)
        
        # Row 2: Employee ID and Calendar Days
        ws.append([
            "EMP ID:", 
            employee.employee_id, 
            "CALENDAR DAYS:", 
            num_days, 
            "Days"
        ] + days)
        
        # Row 3: Employee Name and Shifts
        ws.append([
            "EMP Name:", 
            employee.employee_name, 
            "PAID DAYS:", 
            "", 
            "Shift"
        ] + attendance_data['shifts'])
        
        # Row 4: Company and Status
        ws.append([
            "COMPANY:", 
            employee.company.name if employee.company else '', 
            "PRESENT DAYS:", 
            attendance_data['total_present_days'], 
            "Status"
        ] + attendance_data['shift_status'])
        
        # Additional rows
        ws.append([
            "LOCATION:", 
            employee.location.name if employee.location else '', 
            "WO:", 
            attendance_data['total_WO_days'], 
            "Duty-In"
        ] + attendance_data['first_logtime'])
        
        ws.append([
            "DEPARTMENT:", 
            employee.department.name if employee.department else '', 
            "WW:", 
            attendance_data['total_WW_days'], 
            "Duty-Out"
        ] + attendance_data['last_logtime'])
        
        # Add remaining rows with attendance data
        self._add_remaining_rows(ws, employee, attendance_data, num_days)

    def _add_remaining_rows(self, ws, employee, attendance_data: Dict, num_days: int):
        """Helper method to add remaining attendance data rows."""
        empty_days = [""] * num_days
        
        rows_data = [
            ["DESIGNATION:", employee.designation.name if employee.designation else '', 
             "FS:", "0", "Duty Hours"] + attendance_data['total_hours'],
            ["EMP TYPE:", employee.job_type if employee.job_type else '', 
             "PH:", "0", "Lunch-Out"] + empty_days,
            ["REPORTING MNG:", 
             f"{employee.reporting_manager.employee_name} [{employee.reporting_manager.employee_id}]" 
             if employee.reporting_manager else '', 
             "CO:", "0", "Lunch-In"] + empty_days,
            ["CL:", "0", "CW:", "0", "Lunch Hours"] + empty_days,
            ["EL:", "0", "LATE ENTRY HOURS:", 
             str(attendance_data['total_late_entry_time']), "OT-In"],
            ["SL:", "0", "EARLY EXIT HOURS:", 
             str(attendance_data['total_early_exit_time']), "OT-Out"],
            ["MEDICAL:", "0", "OT HOURS:", 
             str(attendance_data['total_OT_time']), "OT Hours"] + 
             attendance_data['overtime'],
            ["ON DUTY:", "0", "LUNCH HOURS:", "", "Late Entry"] + 
             attendance_data['late_entries'],
            [self._get_shift_summary(attendance_data['shifts']), "", 
             "WORKING HOURS:", str(attendance_data['total_working_time']), 
             "Early Exit"] + attendance_data['early_exits']
        ]
        
        for row_data in rows_data:
            ws.append(row_data)

    def _get_shift_summary(self, shifts: List[str]) -> str:
        """Generate shift summary string."""
        return (f"GS:{shifts.count('GS')} OS:{shifts.count('OS')} "
                f"FS:{shifts.count('FS')} SS:{shifts.count('SS')} "
                f"NS:{shifts.count('NS')}")

    def apply_cell_styles(self, ws, start_row: int, num_days: int):
        """
        Apply styles to cells in the worksheet.
        
        Args:
            ws: Worksheet object
            start_row: Starting row number
            num_days: Number of days in month
        """
        # Apply header style
        header_cell = ws.cell(row=start_row, column=1)
        header_cell.font = Font(bold=True)
        header_cell.border = self.THIN_BORDER

        # Apply styles to data cells
        for row in range(start_row + 1, start_row + 15):
            for col in range(1, num_days + 6):
                cell = ws.cell(row=row, column=col)
                cell.border = self.THIN_BORDER

                # Style header column
                if col == 1 or col == 3 or col == 5:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="E3F2FD",
                        end_color="E3F2FD",
                        fill_type="solid"
                    )
                
                if col == 2 or col == 4:
                    cell.font = Font(bold=True)

                # Style status cells
                if row == start_row + 3 and col > 5:
                    status = cell.value
                    if status in self.SHIFT_STATUS_STYLES:
                        style_dict = self.SHIFT_STATUS_STYLES[status]
                        cell.fill = style_dict['fill']
                        cell.font = style_dict['font']
                        cell.border = style_dict['border']
                        cell.alignment = style_dict['alignment']

                # Style days row
                if row == start_row + 1 and col > 5:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="F5F5F5",
                        end_color="F5F5F5",
                        fill_type="solid"
                    )

    @staticmethod
    def get_last_day_of_month(first_day: date) -> date:
        """Get the last day of the month."""
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def auto_adjust_column_width(self, ws, min_width: int = 10):
        """
        Adjust column widths based on content, excluding header rows and columns beyond 5.
        Only adjusts the first 5 columns (employee details columns).
        
        Args:
            ws: Worksheet object
            min_width: Minimum column width to apply
        """
        header_rows = getattr(ws, '_header_rows', set())
        
        # Only process first 5 columns
        for col_idx in range(1, 6):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            column = ws[column_letter]
            
            for cell in column:
                # Skip cells in header rows
                if cell.row in header_rows:
                    continue
                
                try:
                    if cell.value:
                        max_length = max(
                            max_length, 
                            len(str(cell.value))
                        )
                except:
                    continue
            
            adjusted_width = max(max_length + 2, min_width)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set fixed width for remaining columns (attendance data)
        for col_idx in range(6, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = min_width

class ExportAllEmployeeAttendanceExcelView2(View):
    """
    View to generate and export detailed attendance Excel report for all employees.
    Includes comprehensive attendance data with styled formatting.
    """

    @property
    def SHIFT_STATUS_STYLES(self):
        """Define and register named styles for shift statuses."""
        if not hasattr(self, '_shift_status_styles'):
            self._shift_status_styles = {
                'P': {
                    'fill': PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
                    'font': Font(color="256029", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'WW': {
                    'fill': PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
                    'font': Font(color="256029", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'A': {
                    'fill': PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
                    'font': Font(color="C63737", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'HD': {
                    'fill': PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"),
                    'font': Font(color="8A5340", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'WO': {
                    'fill': PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"),
                    'font': Font(color="8A5340", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'PH': {
                    'fill': PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"),
                    'font': Font(color="8A5340", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'FH': {
                    'fill': PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"),
                    'font': Font(color="8A5340", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'MP': {
                    'fill': PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),
                    'font': Font(color="D84315", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                },
                'IH': {
                    'fill': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
                    'font': Font(color="856404", bold=True),
                    'border': self.THIN_BORDER,
                    'alignment': Alignment(horizontal='center')
                }
            }
        return self._shift_status_styles

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """Handle GET request to generate the Excel file."""
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse(
                {'error': 'Month and year are required parameters.'}, 
                status=400
            )

        # Initialize date parameters
        first_day = date(int(year), int(month), 1)
        last_day = self.get_last_day_of_month(first_day)
        num_days = last_day.day

        # Fetch all required data at once
        attendance_data, employee_data = self.fetch_data(month, year)

        # Generate Excel
        wb = self.generate_excel(
            attendance_data, 
            employee_data, 
            first_day, 
            last_day, 
            num_days
        )

        # Prepare response
        month_name = first_day.strftime('%B')
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename=All_Employee_Attendance_{month_name}_{year}.xlsx'
        )
        wb.save(response)
        return response

    def fetch_data(self, month: str, year: str) -> Tuple[Dict, Dict]:
        """
        Fetch all required attendance and employee data efficiently.
        """
        attendance_records = (
            Attendance.objects.filter(
                logdate__year=year, 
                logdate__month=month
            )
            .select_related('employeeid')
            .order_by('employeeid', 'logdate')
        )
        
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        employees = (
            Employee.objects.all()
            .select_related(
                'department', 
                'designation', 
                'location', 
                'company', 
                'reporting_manager'
            )
        )
        
        employee_data = {
            emp.id: emp for emp in employees
        }

        return attendance_data, employee_data

    def generate_excel(
        self, 
        attendance_data: Dict, 
        employee_data: Dict, 
        first_day: date, 
        last_day: date, 
        num_days: int
    ) -> Workbook:
        """Generate Excel workbook with attendance data."""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{first_day.month}_{first_day.year}"

        row_num = 1
        for emp_id, employee in employee_data.items():
            if emp_id in attendance_data:
                self.add_employee_attendance(
                    ws,
                    employee,
                    attendance_data[emp_id],
                    first_day,
                    last_day,
                    num_days,
                    row_num
                )
                row_num += 12

        self.auto_adjust_column_width(ws)
        return wb

    def add_employee_attendance(
        self, 
        ws, 
        employee, 
        attendance_records: List, 
        first_day: date, 
        last_day: date, 
        num_days: int, 
        start_row: int
    ):
        """Add employee attendance data to worksheet with proper formatting."""
        attendance_data = self.process_attendance_records(
            attendance_records, 
            num_days, 
            first_day
        )
        
        # Add employee details rows
        self.add_employee_details_rows(
            ws, 
            employee, 
            attendance_data, 
            num_days, 
            start_row,
            first_day
        )

        # Apply styles
        self.apply_cell_styles(ws, start_row, num_days)

    def process_attendance_records(
        self, 
        records: List, 
        num_days: int, 
        first_day: date
    ) -> Dict:
        """
        Process attendance records into required format.
        
        Args:
            records: List of attendance records
            num_days: Number of days in the month
            first_day: First day of the month
            
        Returns:
            Dict containing processed attendance data
        """
        # Initialize data structures
        attendance_map = {
            record.logdate.day: record 
            for record in records
        }
        
        # Initialize attendance data
        data = {
            'shifts': [],
            'shift_status': [],
            'first_logtime': [],
            'last_logtime': [],
            'total_hours': [],
            'late_entries': [],
            'early_exits': [],
            'overtime': []
        }
        
        # Process each day
        for day in range(1, num_days + 1):
            record = attendance_map.get(day)
            
            if record:
                data['shifts'].append(record.shift or "")
                data['shift_status'].append(record.shift_status or "")
                data['first_logtime'].append(record.first_logtime or "")
                data['last_logtime'].append(record.last_logtime or "")
                data['total_hours'].append(record.total_time or "")
                data['late_entries'].append(record.late_entry or "")
                data['early_exits'].append(record.early_exit or "")
                data['overtime'].append(record.overtime or "")
            else:
                # Add empty values for missing records
                for key in data:
                    data[key].append("")
        
        # Calculate totals
        data['total_present_days'] = (
            data['shift_status'].count('P') +
            data['shift_status'].count('IH') +
            data['shift_status'].count('HD') 
        )

        data['total_absent_days'] = (
            data['shift_status'].count('A') 
        )
        data['total_WO_days'] = data['shift_status'].count('WO') + data['shift_status'].count('WW')
        data['total_WW_days'] = data['shift_status'].count('WW')
        
        # Calculate time totals
        data['total_late_entry_time'] = sum(
            [t for t in data['late_entries'] if t], 
            timedelta()
        )
        data['total_early_exit_time'] = sum(
            [t for t in data['early_exits'] if t], 
            timedelta()
        )
        data['total_OT_time'] = sum(
            [t for t in data['overtime'] if t], 
            timedelta()
        )
        data['total_working_time'] = sum(
            [t for t in data['total_hours'] if t], 
            timedelta()
        )
        
        return data

    def add_employee_details_rows(
        self, 
        ws, 
        employee, 
        attendance_data: Dict, 
        num_days: int, 
        start_row: int,
        first_day: date
    ):
        """
        Add employee details and attendance rows to worksheet.
        
        Args:
            ws: Worksheet object
            employee: Employee object
            attendance_data: Processed attendance data
            num_days: Number of days in month
            start_row: Starting row number
        """
        days = [str(day) for day in range(1, num_days + 1)]

        first_day_of_month = date(first_day.year, first_day.month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        # Row 1: Employee Header - Stored separately for width adjustment exclusion
        header_text_1 = (f"Monthly Movements Register for the period of: {first_day_of_month} to {last_day_of_month} | Printed Date: {datetime.now().strftime('%d-%m-%Y')}")
        header_text_2 = (
            f"EMP ID: {employee.employee_id} | EMP Name: {employee.employee_name} | "
            f"Department: {employee.department.name if employee.department else ''} | "
            f"Designation: {employee.designation.name if employee.designation else ''} | "
            f"Type: {employee.job_type if employee.job_type else ''} | Location: {employee.location.name if employee.location else ''} | "
            f"Company: {employee.company.name if employee.company else ''}"
        )
        
        # Store header information for width adjustment exclusion
        if not hasattr(ws, '_header_rows'):
            ws._header_rows = set()
        ws._header_rows.add(start_row)
        ws._header_rows.add(start_row + 1)
        
        # Merge and set header
        ws.merge_cells(
            start_row=start_row, 
            start_column=1, 
            end_row=start_row, 
            end_column=num_days + 5
        )

        ws.merge_cells(
            start_row=start_row + 1, 
            start_column=1, 
            end_row=start_row + 1, 
            end_column=num_days + 5
        )

        header_cell = ws.cell(row=start_row, column=1, value=header_text_1)
        # header_cell.font = Font(bold=True)

        header_cell = ws.cell(row=start_row + 1, column=1, value=header_text_2)
        header_cell.font = Font(bold=True)
        
        # Row 2: Employee ID and Calendar Days
        ws.append([
            "", 
            # employee.employee_id, 
            "Days"
        ] + days)
        
        # Row 3: Employee Name and Shifts
        ws.append([
            "EMP ID:",
            "Shift"
        ] + attendance_data['shifts'])
        
        # Row 4: Company and Status
        ws.append([
            employee.employee_id,
            "Status"
        ] + attendance_data['shift_status'])
        
        # Additional rows
        ws.append([
            f"Present: {attendance_data['total_present_days']}",
            "Duty-In"
        ] + attendance_data['first_logtime'])
        
        ws.append([
            f"Absent: {attendance_data['total_absent_days']}",
            "Duty-Out"
        ] + attendance_data['last_logtime'])
        
        # Add remaining rows with attendance data
        self._add_remaining_rows(ws, employee, attendance_data, num_days)

    def _add_remaining_rows(self, ws, employee, attendance_data: Dict, num_days: int):
        """Helper method to add remaining attendance data rows."""
        empty_days = [""] * num_days
        
        rows_data = [
            [f"Weekly Off: {attendance_data['total_WO_days'] + attendance_data['total_WW_days']}","Duty Hours"] + attendance_data['total_hours'],
            [f"OT HOURS: {str(attendance_data['total_OT_time'])}", "OT Hours"] + attendance_data['overtime'],
            [f"Late Entry: {str(attendance_data['total_late_entry_time'])}", "Late Entry"] + attendance_data['late_entries'],
            [f"Early Exit: {str(attendance_data['total_early_exit_time'])}", "Early Exit"] + attendance_data['early_exits']
        ]
        
        for row_data in rows_data:
            ws.append(row_data)

    def _get_shift_summary(self, shifts: List[str]) -> str:
        """Generate shift summary string."""
        return (f"GS:{shifts.count('GS')} OS:{shifts.count('OS')} "
                f"FS:{shifts.count('FS')} SS:{shifts.count('SS')} "
                f"NS:{shifts.count('NS')}")

    def apply_cell_styles(self, ws, start_row: int, num_days: int):
        """
        Apply styles to cells in the worksheet.
        
        Args:
            ws: Worksheet object
            start_row: Starting row number
            num_days: Number of days in month
        """
        # Apply header style
        header_cell = ws.cell(row=start_row, column=1)
        header_cell.font = Font(bold=True)
        # header_cell.border = self.THIN_BORDER

        # Apply styles to data cells
        for row in range(start_row + 2, start_row + 11):
            for col in range(1, num_days + 6):
                cell = ws.cell(row=row, column=col)
                cell.border = self.THIN_BORDER

                # Style header column
                if col == 2:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="E3F2FD",
                        end_color="E3F2FD",
                        fill_type="solid"
                    )
                    cell.font = Font(bold=True)
                
                if col == 1:
                    cell.font = Font(bold=True)

                # Style status cells
                if row == start_row + 4 and col > 2:
                    status = cell.value
                    if status in self.SHIFT_STATUS_STYLES:
                        style_dict = self.SHIFT_STATUS_STYLES[status]
                        cell.fill = style_dict['fill']
                        cell.font = style_dict['font']
                        cell.border = style_dict['border']
                        cell.alignment = style_dict['alignment']

                # Style days row
                if row == start_row + 2 and col > 2:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="F5F5F5",
                        end_color="F5F5F5",
                        fill_type="solid"
                    )

    @staticmethod
    def get_last_day_of_month(first_day: date) -> date:
        """Get the last day of the month."""
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def auto_adjust_column_width(self, ws, min_width: int = 10):
        """
        Adjust column widths based on content, excluding header rows and columns beyond 5.
        Only adjusts the first 5 columns (employee details columns).
        
        Args:
            ws: Worksheet object
            min_width: Minimum column width to apply
        """
        header_rows = getattr(ws, '_header_rows', set())
        
        # Only process first 5 columns
        for col_idx in range(1, 6):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            column = ws[column_letter]
            
            for cell in column:
                # Skip cells in header rows
                if cell.row in header_rows:
                    continue
                
                try:
                    if cell.value:
                        max_length = max(
                            max_length, 
                            len(str(cell.value))
                        )
                except:
                    continue
            
            adjusted_width = max(max_length + 2, min_width)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set fixed width for remaining columns (attendance data)
        for col_idx in range(6, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = min_width
        
class LastLogIdView(APIView):
    def get(self, request):
        last_log = LastLogId.objects.first()  # or .last() if you want the last record
        serializer = serializers.LastLogIdSerializer(last_log)
        return Response(serializer.data)
    
class MandaysAttendanceListCreate(generics.ListCreateAPIView):
    """
    API view for listing and creating mandays attendance records.
    """
    queryset = ManDaysAttendance.objects.order_by('-logdate').all()
    serializer_class = serializers.ManDaysAttendanceSerializer
    pagination_class = DefaultPagination
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    filter_fields = '__all__'
    search_fields = ['employee_id', 'employee_name', 'device_enroll_id', 'logdate', 'shift', 
                    '=shift_status', 'first_logtime', 'last_logtime', 'total_time', 'late_entry', 
                    'early_exit', 'overtime', 'company_name', 'location_name']

    def get_queryset(self):
        """
        Get the queryset for listing attendance records with optional search query.
        """
        search_query = self.request.GET.get('search')
        date_query = self.request.GET.get('date')

        # Get date range parameters
        date_from = self.request.GET.get('date_from') 
        date_to = self.request.GET.get('date_to') 
        employee_id = self.request.GET.get('employeeid')  
        employee_ids = self.request.GET.get('employee_ids')
        company_names = self.request.GET.get('company_name')
        location_names = self.request.GET.get('location_name') 
        department_names = self.request.GET.get('department_name')
        designation_names = self.request.GET.get('designation_name')

        queryset = ManDaysAttendance.objects.order_by('-logdate').all()

        if search_query:
            queryset = queryset.filter(
                Q(employeeid__employee_id__iexact=search_query) |
                Q(employeeid__device_enroll_id__iexact=search_query) |
                Q(logdate__iexact=search_query) |
                Q(employeeid__company__name__iexact=search_query) |
                Q(employeeid__location__name__iexact=search_query) |
                Q(employeeid__department__name__iexact=search_query) |
                Q(employeeid__designation__name__iexact=search_query) 
            )

        if date_query:
            try:
                # Convert date_query to a datetime object
                date_obj = datetime.strptime(date_query, '%m-%d-%Y').date()
                queryset = queryset.filter(logdate=date_obj)
            except ValueError:
                # Handle invalid date format
                pass

        # Filter by date range if both date_from and date_to are provided
        if date_from and date_to:
            try:
                date_from_obj = datetime.strptime(date_from, '%m-%d-%Y').date()
                date_to_obj = datetime.strptime(date_to, '%m-%d-%Y').date()
                queryset = queryset.filter(logdate__range=[date_from_obj, date_to_obj]).order_by('logdate')
            except ValueError:
                pass 

        if employee_id:
            queryset = queryset.filter(employeeid=employee_id)

        if employee_ids:
            employee_ids_list = [id.strip() for id in employee_ids.split(',')]
            queryset = queryset.filter(employeeid__employee_id__in=employee_ids_list)

        if company_names:
            company_names_list = [name.strip() for name in company_names.split(',')]
            queryset = queryset.filter(employeeid__company__name__in=company_names_list)
        
        if location_names:
            location_names_list = [name.strip() for name in location_names.split(',')]
            queryset = queryset.filter(employeeid__location__name__in=location_names_list)

        if department_names:
            department_names_list = [name.strip() for name in department_names.split(',')]
            queryset = queryset.filter(employeeid__department__name__in=department_names_list)

        if designation_names:
            designation_names_list = [name.strip() for name in designation_names.split(',')]
            queryset = queryset.filter(employeeid__designation__name__in=designation_names_list)

        return queryset
    
    def get(self, request, *args, **kwargs):
        """
        Get the list of attendance records with optional search query.
        """
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class ManDaysAttendanceExcelExport(View):
    HEADERS = (
        "Employee ID", "Device Enroll ID", "Employee Name", "Company", "Location", 
        "Department", "Designation", "Employee Type",
        "Log Date", "Duty In 1", "Duty Out 1", "Total Hours", "Duty In 2", 
        "Duty Out 2", "Total Hours", "Duty In 3", "Duty Out 3", "Total Hours",
        "Duty In 4", "Duty Out 4", "Total Hours", "Duty In 5", "Duty Out 5",
        "Total Hours", "Duty In 6", "Duty Out 6", "Total Hours", "Duty In 7",
        "Duty Out 7", "Total Hours", "Duty In 8", "Duty Out 8", "Total Hours",
        "Duty In 9", "Duty Out 9", "Total Hours", "Duty In 10", "Duty Out 10",
        "Mandays Worked Hours"
    )

    def get_queryset(self, request):
        """Get filtered queryset with all related fields"""
        employee_id = request.GET.get('employee_id')
        date_str = request.GET.get('date')
        month = request.GET.get('month')
        year = request.GET.get('year')

        date_from = self.request.GET.get('date_from') 
        date_to = self.request.GET.get('date_to')   
        employee_ids = self.request.GET.get('employee_ids')
        company_names = self.request.GET.get('company_name')
        location_names = self.request.GET.get('location_name') 
        department_names = self.request.GET.get('department_name')
        designation_names = self.request.GET.get('designation_name')

        # Use select_related for foreign key relationships
        queryset = ManDaysAttendance.objects.select_related(
            'employeeid',
            'employeeid__company',
            'employeeid__location',
            'employeeid__department',
            'employeeid__designation',
        ).order_by('-logdate')

        # Date range filter
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        if date_from and date_to:
            try:
                date_from_obj = datetime.strptime(date_from, '%m-%d-%Y').date()
                date_to_obj = datetime.strptime(date_to, '%m-%d-%Y').date()
                queryset = queryset.filter(logdate__range=[date_from_obj, date_to_obj])
            except ValueError:
                pass
        
        employee_ids = request.GET.get('employee_ids')
        if employee_ids:
            employee_ids_list = [id.strip() for id in employee_ids.split(',')]
            queryset = queryset.filter(employeeid__employee_id__in=employee_ids_list)

        # Apply filters
        if employee_id:
            queryset = queryset.filter(Q(employeeid__employee_id__iexact=employee_id))
        if date_str:
            queryset = queryset.filter(logdate=date_str)
        if month:
            queryset = queryset.filter(logdate__month=month)
        if year:
            queryset = queryset.filter(logdate__year=year)

        if employee_ids:
            employee_ids_list = [id.strip() for id in employee_ids.split(',')]
            queryset = queryset.filter(employeeid__employee_id__in=employee_ids_list)

        if company_names:
            company_names_list = [name.strip() for name in company_names.split(',')]
            queryset = queryset.filter(employeeid__company__name__in=company_names_list)
        
        if location_names:
            location_names_list = [name.strip() for name in location_names.split(',')]
            queryset = queryset.filter(employeeid__location__name__in=location_names_list)

        if department_names:
            department_names_list = [name.strip() for name in department_names.split(',')]
            queryset = queryset.filter(employeeid__department__name__in=department_names_list)

        if designation_names:
            designation_names_list = [name.strip() for name in designation_names.split(',')]
            queryset = queryset.filter(employeeid__designation__name__in=designation_names_list)

        return queryset

    def format_timedelta(self, td):
        """Format timedelta or return empty string"""
        return td if td and td != timedelta(0) else ""

    def setup_worksheet(self, wb):
        """Setup worksheet with headers and styling"""
        ws = wb.active
        ws.title = "Mandays Attendance Report"
        
        # Cache styles
        header_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Write headers
        for col_num, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[cell.column_letter].width = len(header) + 7

        ws.freeze_panes = 'A2'
        return ws

    def get_record_data(self, record):
        """Extract data from record into tuple"""
        return (
            record.employeeid.employee_id,
            record.employeeid.device_enroll_id,
            record.employeeid.employee_name,
            record.employeeid.company.name,
            record.employeeid.location.name,
            record.employeeid.department.name if record.employeeid.department else "",
            record.employeeid.designation.name if record.employeeid.designation else "",
            record.employeeid.job_type if record.employeeid.job_type else "",
            record.logdate,
            record.duty_in_1,
            record.duty_out_1,
            self.format_timedelta(record.total_time_1),
            record.duty_in_2,
            record.duty_out_2,
            self.format_timedelta(record.total_time_2),
            record.duty_in_3,
            record.duty_out_3,
            self.format_timedelta(record.total_time_3),
            record.duty_in_4,
            record.duty_out_4,
            self.format_timedelta(record.total_time_4),
            record.duty_in_5,
            record.duty_out_5,
            self.format_timedelta(record.total_time_5),
            record.duty_in_6,
            record.duty_out_6,
            self.format_timedelta(record.total_time_6),
            record.duty_in_7,
            record.duty_out_7,
            self.format_timedelta(record.total_time_7),
            record.duty_in_8,
            record.duty_out_8,
            self.format_timedelta(record.total_time_8),
            record.duty_in_9,
            record.duty_out_9,
            self.format_timedelta(record.total_time_9),
            record.duty_in_10,
            record.duty_out_10,
            self.format_timedelta(record.total_hours_worked)
        )

    def get(self, request, *args, **kwargs):
        # Get filtered queryset
        queryset = self.get_queryset(request)
        
        # Convert queryset to tuple of tuples for better performance
        records = tuple(
            self.get_record_data(record) for record in queryset
        )

        # Create workbook and setup worksheet
        wb = openpyxl.Workbook()
        ws = self.setup_worksheet(wb)
        
        # Cache alignment style
        center_alignment = Alignment(horizontal='center')

        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write data efficiently
        for row_num, record_data in enumerate(records, 2):
            for col_num, value in enumerate(record_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = center_alignment
                cell.border = thin_border

        # Create response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=Mandays_Attendance_Report.xlsx"
        wb.save(response)

        return response      
    
class ManDaysWorkedExcelExport(View):
    """
    API view for exporting the mandays worked data to an Excel file.
    """

    def get_duty_times(self, record):
        """Helper method to get first duty in and last duty out times."""
        # Get all duty in times
        duty_in_times = [
            getattr(record, f'duty_in_{i}')
            for i in range(1, 11)
            if getattr(record, f'duty_in_{i}') is not None
        ]
        
        # Get all duty out times
        duty_out_times = [
            getattr(record, f'duty_out_{i}')
            for i in range(1, 11)
            if getattr(record, f'duty_out_{i}') is not None
        ]
        
        first_duty_in = duty_in_times[0] if duty_in_times else None
        last_duty_out = duty_out_times[-1] if duty_out_times else None
        
        return first_duty_in, last_duty_out

    def get(self, request, *args, **kwargs):
        employee_id = request.GET.get('employee_id')
        date_str = request.GET.get('date')
        month = request.GET.get('month')
        year = request.GET.get('year')

        date_from = self.request.GET.get('date_from') 
        date_to = self.request.GET.get('date_to')
        employee_ids = self.request.GET.get('employee_ids')
        company_names = self.request.GET.get('company_name')
        location_names = self.request.GET.get('location_name') 
        department_names = self.request.GET.get('department_name')
        designation_names = self.request.GET.get('designation_name')

        queryset = ManDaysAttendance.objects.order_by('-logdate').all()

        # Date range filter
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        if date_from and date_to:
            try:
                date_from_obj = datetime.strptime(date_from, '%m-%d-%Y').date()
                date_to_obj = datetime.strptime(date_to, '%m-%d-%Y').date()
                queryset = queryset.filter(logdate__range=[date_from_obj, date_to_obj])
            except ValueError:
                pass
        
        employee_ids = request.GET.get('employee_ids')
        if employee_ids:
            employee_ids_list = [id.strip() for id in employee_ids.split(',')]
            queryset = queryset.filter(employeeid__employee_id__in=employee_ids_list)

        if employee_id:
            queryset = queryset.filter(Q(employeeid__employee_id__iexact=employee_id))
        if date_str:
            queryset = queryset.filter(logdate=date_str)
        if month:
            queryset = queryset.filter(logdate__month=month)
        if year:
            queryset = queryset.filter(logdate__year=year)

        if employee_ids:
            employee_ids_list = [id.strip() for id in employee_ids.split(',')]
            queryset = queryset.filter(employeeid__employee_id__in=employee_ids_list)

        if company_names:
            company_names_list = [name.strip() for name in company_names.split(',')]
            queryset = queryset.filter(employeeid__company__name__in=company_names_list)
        
        if location_names:
            location_names_list = [name.strip() for name in location_names.split(',')]
            queryset = queryset.filter(employeeid__location__name__in=location_names_list)

        if department_names:
            department_names_list = [name.strip() for name in department_names.split(',')]
            queryset = queryset.filter(employeeid__department__name__in=department_names_list)

        if designation_names:
            designation_names_list = [name.strip() for name in designation_names.split(',')]
            queryset = queryset.filter(employeeid__designation__name__in=designation_names_list)        

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mandays Worked Report"

        headers = ["Employee ID", "Device Enroll ID", "Employee Name", "Company", "Location", "Jobtype", "Department", "Employee Type", "Designation", "Log Date", "Duty In First", "Duty Out Last", "Mandays Worked Hours"]

        row_num = 1

        # Set font style and background color for headers
        header_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[ws.cell(row=row_num, column=col_num).column_letter].width = len(header) + 7
        ws.freeze_panes = 'A2'

        for row_num, record in enumerate(queryset, 2):
            first_duty_in, last_duty_out = self.get_duty_times(record)

            ws.cell(row=row_num, column=1, value=record.employeeid.employee_id)
            ws.cell(row=row_num, column=2, value=record.employeeid.device_enroll_id)
            ws.cell(row=row_num, column=3, value=record.employeeid.employee_name)
            ws.cell(row=row_num, column=4, value=record.employeeid.company.name)
            ws.cell(row=row_num, column=5, value=record.employeeid.location.name)
            ws.cell(row=row_num, column=6, value=record.employeeid.job_type)
            if record.employeeid.department is not None:
                ws.cell(row=row_num, column=7, value=record.employeeid.department.name)
            else:
                ws.cell(row=row_num, column=7, value="")
            ws.cell(row=row_num, column=8, value=record.employeeid.category if record.employeeid.category else "")
            if record.employeeid.designation is not None:
                ws.cell(row=row_num, column=9, value=record.employeeid.designation.name)
            else:
                ws.cell(row=row_num, column=9, value="")
            ws.cell(row=row_num, column=10, value=record.logdate)
            ws.cell(row=row_num, column=11, value=first_duty_in)
            ws.cell(row=row_num, column=12, value=last_duty_out)
            ws.cell(row=row_num, column=13, value=record.total_hours_worked if record.total_hours_worked else "")

            cell.alignment = Alignment(horizontal='center')

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Mandays_Worked_Report.xlsx"
        wb.save(response)

        return response  
    
class ManDaysMissedPunchExcelExport(View):

    def get(self, request, *args, **kwargs):
        employee_id = request.GET.get('employee_id')
        date_str = request.GET.get('date')
        month = request.GET.get('month')
        year = request.GET.get('year')

        queryset = ManDaysMissedPunchAttendance.objects.order_by('-logdate').all()

        if employee_id:
            queryset = queryset.filter(Q(employeeid__employee_id__iexact=employee_id))
        if date_str:
            queryset = queryset.filter(logdate=date_str)
        if month:
            queryset = queryset.filter(logdate__month=month)
        if year:
            queryset = queryset.filter(logdate__year=year)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mandays Attendance Report"

        headers = ["Employee ID", "Device Enroll ID", "Employee Name", "Company", "Location", "Log Date", "Duty In 1", "Duty Out 1", 
                   "Duty In 2", "Duty Out 2", "Duty In 3", "Duty Out 3", "Duty In 4", "Duty Out 4", 
                   "Duty In 5", "Duty Out 5", "Duty In 6", "Duty Out 6", "Duty In 7", "Duty Out 7", 
                   "Duty In 8", "Duty Out 8", "Duty In 9", "Duty Out 9", "Duty In 10", "Duty Out 10"]
        
        row_num = 1

        # Set font style and background color for headers
        header_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[ws.cell(row=row_num, column=col_num).column_letter].width = len(header) + 7
        ws.freeze_panes = 'A2'

        for row_num, record in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=record.employeeid.employee_id)
            ws.cell(row=row_num, column=2, value=record.employeeid.device_enroll_id)
            ws.cell(row=row_num, column=3, value=record.employeeid.employee_name)
            ws.cell(row=row_num, column=4, value=record.employeeid.company.name)
            ws.cell(row=row_num, column=5, value=record.employeeid.location.name)
            ws.cell(row=row_num, column=6, value=record.logdate)
            ws.cell(row=row_num, column=7, value=record.duty_in_1)
            ws.cell(row=row_num, column=8, value=record.duty_out_1)
            ws.cell(row=row_num, column=10, value=record.duty_in_2)
            ws.cell(row=row_num, column=11, value=record.duty_out_2)
            ws.cell(row=row_num, column=13, value=record.duty_in_3)
            ws.cell(row=row_num, column=14, value=record.duty_out_3)
            ws.cell(row=row_num, column=16, value=record.duty_in_4)
            ws.cell(row=row_num, column=17, value=record.duty_out_4)
            ws.cell(row=row_num, column=19, value=record.duty_in_5)
            ws.cell(row=row_num, column=20, value=record.duty_out_5)
            ws.cell(row=row_num, column=22, value=record.duty_in_6)
            ws.cell(row=row_num, column=23, value=record.duty_out_6)
            ws.cell(row=row_num, column=25, value=record.duty_in_7)
            ws.cell(row=row_num, column=26, value=record.duty_out_7)
            ws.cell(row=row_num, column=28, value=record.duty_in_8)
            ws.cell(row=row_num, column=29, value=record.duty_out_8)
            ws.cell(row=row_num, column=31, value=record.duty_in_9)
            ws.cell(row=row_num, column=32, value=record.duty_out_9)
            ws.cell(row=row_num, column=34, value=record.duty_in_10)
            ws.cell(row=row_num, column=35, value=record.duty_out_10)

            cell.alignment = Alignment(horizontal='center')

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Mandays_Missed_Punch_Report.xlsx"
        wb.save(response)

        return response      

class ExportLogsExcelView(View):
    def get_filtered_queryset(self, request):
        """
        Get the filtered queryset using the same logic as LogsListCreate
        """
        queryset = Logs.objects.all()

        # Get query parameters
        search_query = request.GET.get('search')
        date_query = request.GET.get('date')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        employee_id = request.GET.get('employeeid')
        employee_ids = request.GET.get('employee_ids')
        direction = request.GET.get('direction')
        
        # Apply search filter (case-insensitive)
        if search_query:
            queryset = queryset.filter(
                Q(employeeid__icontains=search_query) |
                Q(shortname__icontains=search_query)
            )

        # Filter by specific date
        if date_query:
            try:
                date_obj = datetime.strptime(date_query, '%Y-%m-%d').date()
                queryset = queryset.filter(log_datetime__date=date_obj)
            except ValueError:
                pass

        # Filter by date range
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(log_datetime__date__gte=date_from_obj)
            except ValueError:
                pass

        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(log_datetime__date__lte=date_to_obj)
            except ValueError:
                pass

        # Filter by single employee ID
        if employee_id:
            queryset = queryset.filter(employeeid=employee_id)

        # Filter by multiple employee IDs
        if employee_ids:
            try:
                employee_ids_list = [id.strip() for id in employee_ids.split(',')]
                queryset = queryset.filter(employeeid__in=employee_ids_list)
            except Exception:
                pass

        # Filter by direction
        if direction:
            queryset = queryset.filter(direction=direction)

        return queryset.order_by('-log_datetime')

    def get(self, request, *args, **kwargs):
        # Get filtered queryset
        queryset = self.get_filtered_queryset(request)
        
        # Fetch only the exact fields we need directly
        logs_data = queryset.values_list(
            'id',
            'employeeid',  # Direct field access since it's a CharField
            'direction',
            'log_datetime',
            'shortname',
            'serialno'
        )

        # Convert to tuple for performance optimization
        logs_data = tuple(logs_data)

        # Initialize workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Logs"

        # Define headers and styles
        headers = ("ID", "Employee ID", "Direction", "Datetime", "Shortname", "Serialno")
        header_font = Font(size=14, bold=True)
        header_fill = PatternFill(
            start_color="D3D3D3",
            end_color="D3D3D3",
            fill_type="solid"
        )
        center_alignment = Alignment(horizontal='center')

        # Write headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            ws.column_dimensions[cell.column_letter].width = len(header) + 7

        # Write data rows efficiently
        for row_num, record in enumerate(logs_data, 2):
            # Each record is a tuple with exactly the fields we need
            ws.cell(row=row_num, column=1, value=record[0])  # ID
            ws.cell(row=row_num, column=2, value=record[1])  # Employee ID
            ws.cell(row=row_num, column=3, value=record[2])  # Direction
            # Handle timezone for datetime value
            ws.cell(row=row_num, column=4, value=record[3].replace(tzinfo=None) if record[3] else None)
            ws.cell(row=row_num, column=5, value=record[4])  # Shortname
            ws.cell(row=row_num, column=6, value=record[5])  # Serialno

            # Apply center alignment to all cells in the row
            for col in range(1, 5):
                ws.cell(row=row_num, column=col).alignment = center_alignment

        # Prepare and return the response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=Logs.xlsx"
        wb.save(response)

        return response

class ResetMandaysView(generics.GenericAPIView):
    """
    API endpoint to reset and reprocess mandays data.
    Handles:
    1. Scheduler shutdown
    2. Data cleanup for last 7 days or all data if no records exist in range
    3. LastLogIdMandays reset to cutoff day - 1 or 0 if no data exists
    4. Mandays reprocessing
    5. Scheduler restart
    """

    def handle_scheduler(self, action='stop'):
        """
        Safely handle scheduler operations
        Returns True if operation was successful
        """
        try:
            scheduler = get_scheduler()
            if action == 'stop':
                if scheduler and scheduler.running:
                    scheduler.shutdown(wait=True)
                return True
            elif action == 'start':
                if scheduler and not scheduler.running:
                    scheduler.start()
                return True
            return False
        except Exception as e:
            return False

    def cleanup_old_data(self, cutoff_date):
        """
        Clean up ManDaysAttendance records.
        First checks if we have at least 100 days of historical data.
        If we don't have 100 days of data, deletes all records for a fresh start.
        Otherwise, only deletes records from cutoff_date forward.
        Returns tuple of (deleted_count, was_full_cleanup)
        """
        try:
            # Get the date of our oldest record
            oldest_record = ManDaysAttendance.objects.order_by('logdate').first()
            
            if not oldest_record:
                # No records at all - return early with (0, True)
                return (0, True)
                
            # Calculate how many days of history we have
            days_of_history = (datetime.now().date() - oldest_record.logdate).days
            
            # If we have less than 100 days of history, delete everything
            if days_of_history < 100:
                result = ManDaysAttendance.objects.all().delete()
                print(f"Only {days_of_history} days of history found, deleting all records")
                return (result[0], True)
            
            # Otherwise, just delete records from cutoff_date forward
            result = ManDaysAttendance.objects.filter(
                logdate__gte=cutoff_date
            ).delete()
            print(f"Found {days_of_history} days of history, performing normal cleanup")
            return (result[0], False)
            
        except Exception:
            raise

    def update_last_log_id(self, cutoff_date, was_full_cleanup=False):
        """
        Update LastLogIdMandays with log ID from day before cutoff date
        If was_full_cleanup is True or no logs found, resets to 0
        """
        try:
            if was_full_cleanup:
                LastLogIdMandays.objects.all().delete()
                new_last_log = LastLogIdMandays.objects.create(last_log_id=0)
                print("Full cleanup performed, resetting last_log_id to 0")
                return new_last_log.last_log_id

            # Find the last log entry from the day before cutoff date
            day_before_cutoff = cutoff_date - timedelta(days=1)
            last_log_before_cutoff = Logs.objects.filter(
                log_datetime__date__lte=day_before_cutoff
            ).order_by('-id').first()

            LastLogIdMandays.objects.all().delete()
            if last_log_before_cutoff:
                new_last_log = LastLogIdMandays.objects.create(
                    last_log_id=last_log_before_cutoff.id
                )
            else:
                # No logs found, reset to 0
                new_last_log = LastLogIdMandays.objects.create(last_log_id=0)
                print("No logs found before cutoff date, resetting last_log_id to 0")
            
            return new_last_log.last_log_id
        except Exception as e:
            raise

    def post(self, request, *args, **kwargs):
        """Handle POST request to reset mandays data"""
        scheduler_stopped = False
        try:
            # Stop scheduler
            scheduler_stopped = self.handle_scheduler('stop')
            if not scheduler_stopped:
                raise Exception("Failed to stop scheduler")

            with transaction.atomic():
                # Calculate cutoff date (100 days ago)
                cutoff_date = datetime.now().date() - timedelta(days=100)
                
                # Execute cleanup operations
                deleted_count, was_full_cleanup = self.cleanup_old_data(cutoff_date)
                last_log_id = self.update_last_log_id(cutoff_date, was_full_cleanup)
                
                # Run mandays command
                call_command('mandays')
                
                response_data = {
                    'message': 'Successfully reset mandays data and restarted processing',
                    'cutoff_date': cutoff_date.isoformat(),
                    'deleted_records': deleted_count,
                    'last_log_id': last_log_id,
                    'was_full_cleanup': was_full_cleanup,
                    'scheduler_status': 'restarted'
                }
                
                # Restart scheduler
                if not self.handle_scheduler('start'):
                    response_data['scheduler_status'] = 'failed_to_restart'
                
                return Response(response_data, status=status.HTTP_200_OK)
                
        except Exception as e:
            # Ensure scheduler is restarted if it was stopped
            if scheduler_stopped:
                self.handle_scheduler('start')
            
            return Response({
                'error': str(e),
                'message': 'Failed to reset mandays data',
                'scheduler_status': 'restarted_after_error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class OvertimeRoundoffRulesView(generics.ListAPIView):
    queryset = OvertimeRoundoffRules.objects.all()
    serializer_class = serializers.OvertimeRoundoffRulesSerializer
    pagination_class = None

    def list(self, request, *args, **kwargs):
        instance = self.queryset.first()
        if not instance:
            instance = OvertimeRoundoffRules.objects.create()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

class OvertimeRoundoffRulesUpdate(generics.UpdateAPIView):
    queryset = OvertimeRoundoffRules.objects.all()
    serializer_class = serializers.OvertimeRoundoffRulesSerializer
    
    def get_object(self):
        return self.queryset.first()

class test_view(APIView):
    """
    API endpoint to query an entire table and return its size, 
    query execution time, data size (in bytes), and optionally the data.
    """
    def get(self, request):
        # Record the start time
        start_time = time.time()

        # Query the table and convert it to a list of dictionaries
        data = list(Logs.objects.all().values())

        # Calculate query time
        query_time = time.time() - start_time

        # Calculate data size in bytes
        data_size = len(json.dumps(data).encode('utf-8'))

        # Prepare the response
        response = {
            "size": len(data),               # Number of rows
            "query_time": query_time,        # Time taken in seconds
            "data_size": data_size,          # Size of the data in bytes
            "data": data                     # Optional: include queried data
        }

        return Response(response)

class ExportMonthlyDutyHourExcel(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        row_num = 1
        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                self.add_employee_data(ws, details, records, first_day_of_month, num_days, row_num)
                row_num += 8

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Attendance_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def add_employee_data(self, ws, details, records, first_day, num_days, start_row):
        """
        Add employee details and attendance data to the worksheet.
        """
        # Add employee header
        header = (
            f"EMP ID: {details['employee_id']} | EMP Name: {details['employee_name']} | "
            f"Department: {details['department']} | Designation: {details['designation']} | "
            f"Type: {details['job_type']} | Location: {details['location']} | Company: {details['company']}"
        )
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_days)
        cell = ws.cell(row=start_row, column=1, value=header)
        cell.font = Font(bold=True)
        # cell.border = self.THIN_BORDER

        # Process attendance records
        days, shifts, statuses, first_logs, last_logs, total_hours = self.process_records(records, first_day, num_days)

        # Adding the rows to the sheet
        rows = [
            ["Days"] + days,
            ["Shift"] + shifts,
            ["Status"] + statuses,
            ["Duty-In"] + first_logs,
            ["Duty-Out"] + last_logs,
            ["Duty Hours"] + total_hours
        ]

        # Iterate through each row of data and add it to the worksheet
        for i, row_data in enumerate(rows, start=start_row + 1):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.THIN_BORDER
                
                # Style the header column (first column)
                if col == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    ws.column_dimensions[cell.column_letter].width = len(str(value)) + 2  # Add padding
                
                # Style the status cells
                if i == start_row + 3:  # Status row
                    style = self.SHIFT_STATUS_STYLES.get(value)
                    if style:
                        cell.style = style
                
                # Style the days row
                if i == start_row + 1 and col > 1:  # Days row (excluding header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")


    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        shifts, statuses, first_logs, last_logs, total_hours = ([] for _ in range(5))

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                shifts.append(record.shift or "")
                statuses.append(record.shift_status or "")
                first_logs.append(record.first_logtime or "")
                last_logs.append(record.last_logtime or "")
                total_hours.append(record.total_time or "")
            else:
                shifts.append("")
                statuses.append("")
                first_logs.append("")
                last_logs.append("")
                total_hours.append("")

        return days, shifts, statuses, first_logs, last_logs, total_hours

class ExportMonthlyMusterRoleExcel(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        row_num = 1
        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                self.add_employee_data(ws, details, records, first_day_of_month, last_day_of_month, num_days, row_num)
                row_num += 4

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Muster_Role_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def add_employee_data(self, ws, details, records, first_day, last_day, num_days, start_row):
        """
        Add employee details and attendance data to the worksheet.
        """
        # Add employee header
        header = (
            f"Monthly Musterrole Register for the Period of: {first_day} to {last_day}"
        )
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_days)
        cell = ws.cell(row=start_row, column=1, value=header)
        cell.font = Font(bold=True)

        # Process attendance records
        days, statuses, absent_total, present_total = self.process_records(records, first_day, num_days)

        # Define header labels
        header_labels = ["EMP ID", "EMP Name", "Department", "Designation", "Type", "Days"]
        row_data = header_labels + days + ["Present", "Absent"]

        # Adding the rows to the sheet
        rows = [
            row_data,
            [f"{details['employee_id']}", f"{details['employee_name']}", f"{details['department']}", f"{details['designation']}", f"{details['job_type']}", "Status"] + statuses + [present_total, absent_total]
        ]

        # Calculate max width for each column
        max_widths = []
        for col in range(len(header_labels)):
            max_width = max(len(str(header_labels[col])), len(str(rows[1][col])))
            max_widths.append(max_width)

        # Iterate through each row of data and add it to the worksheet
        for i, row_data in enumerate(rows, start=start_row + 1):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.THIN_BORDER
                
                # Style the header column (first columns)
                if col < 7 :
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    # Adjust column width based on max content width
                    if col <= len(max_widths):
                        ws.column_dimensions[cell.column_letter].width = max_widths[col-1] + 2  # Add padding
                
                # Style the "Present" column (last but one)
                if col == len(header_labels) + num_days + 1:  # This is the column for "Present"
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                
                # Style the "Absent" column (last column)
                if col == len(header_labels) + num_days + 2:  # This is the column for "Absent"
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

                # Style the status cells
                if i == start_row + 2:  # Status row
                    style = self.SHIFT_STATUS_STYLES.get(value)
                    if style:
                        cell.style = style
                
                # Style the days row
                if i == start_row + 1 and col > 6:  # Days row (excluding header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]

        statuses = []
        absent_total = 0
        present_total = 0

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                statuses.append(record.shift_status or "")
            else:
                statuses.append("")  # Append an empty string if no record exists
        
        for status_code in statuses:
            if status_code == 'A':
                absent_total += 1
            elif status_code == 'HD':
                present_total += 0.5
                absent_total += 0.5
            elif status_code == 'P':
                present_total += 1
            elif status_code == 'WW':
                present_total += 1
            elif status_code == 'IH':
                present_total += 1

        return days, statuses, absent_total, present_total

class ExportMonthlyMusterRoleExcel2(View):
    """
    View to generate and export an Excel file containing monthly muster role details for all employees.
    """

    # Predefined styles for shift statuses (keeping the same as original)
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"MusterRole_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        # Add main header
        main_header = f"Monthly Muster Role Register for the Period: {first_day_of_month} to {last_day_of_month} | Printed: {datetime.now().strftime('%d-%m-%Y')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 6)
        cell = ws.cell(row=1, column=1, value=main_header)
        cell.font = Font(bold=True)

        # Add column headers
        header_labels = ["EMP ID", "EMP Name", "Company", "Location", "Department", "Designation", "Type", "Days"]
        days_header = [str(day) for day in range(1, num_days + 1)]
        total_header = ["Present", "Absent"]
        full_header = header_labels + days_header + total_header
        ws.append(full_header)

        # Style header row
        for col in range(1, len(full_header) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            cell.border = self.THIN_BORDER

            if col >= 9:  # Style days and total columns
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # Add employee data rows
        row_num = 3  # Start from row 3 after headers
        total_present = 0
        total_absent = 0

        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                days, statuses, absent_total, present_total = self.process_records(records, first_day_of_month, num_days)
                data_row = [
                    details['employee_id'],
                    details['employee_name'],
                    details['company'],
                    details['location'],
                    details['department'],
                    details['designation'],
                    details['job_type'],
                    "Status",
                ] + statuses + [present_total, absent_total]
                ws.append(data_row)

                # Apply styling to data row
                for col in range(1, len(data_row) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = self.THIN_BORDER
                    
                    if col <= 8:  # Style employee info columns
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    if col >= 9 and col < len(data_row) - 1:  # Style status cells
                        status = data_row[col - 1]
                        style = self.SHIFT_STATUS_STYLES.get(status)
                        if style:
                            cell.style = style
                    
                    # Style Present total column
                    if col == len(data_row) - 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    
                    # Style Absent total column
                    if col == len(data_row):
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

                row_num += 1
                total_present += present_total
                total_absent += absent_total

        # Add grand total row
        grand_total_row = [""] * (len(full_header) - 3) + ["Grand Total", total_present, total_absent]
        ws.append(grand_total_row)

        # Style grand total row
        for col in range(len(full_header) - 2, len(full_header) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.border = self.THIN_BORDER

        # Adjust column widths
        self.adjust_column_widths(ws)

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Muster_Role_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'company': emp.company.name,
                'location': emp.location.name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data
    
    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def adjust_column_widths(self, ws):
        """
        Adjust column widths based on the maximum content length in each column,
        skipping merged cells and applying a minimum and maximum width limit.
        """
        max_width = {}
        max_allowed_width = 30  # Set a maximum width limit for columns
        min_allowed_width = 10  # Set a minimum width limit for columns

        for row in ws.iter_rows():
            for cell in row:
                # Skip merged cells as they don't have column_letter
                if isinstance(cell, MergedCell):
                    continue
                column = cell.column_letter
                value_length = len(str(cell.value)) if cell.value is not None else 0
                if column not in max_width or value_length > max_width[column]:
                    max_width[column] = min(value_length, max_allowed_width)  # Apply max width limit

        for col_letter, width in max_width.items():
            # Apply minimum and maximum width, adding padding
            adjusted_width = max(min_allowed_width, width + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

    # Keep existing fetch_data, get_last_day_of_month, adjust_column_widths methods as they are

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        statuses = []
        absent_total = 0
        present_total = 0

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                statuses.append(record.shift_status or "")
            else:
                statuses.append("")
        
        for status_code in statuses:
            if status_code == 'A':
                absent_total += 1
            elif status_code == 'HD':
                present_total += 0.5
                absent_total += 0.5
            elif status_code == 'P':
                present_total += 1
            elif status_code == 'WW':
                present_total += 1
            elif status_code == 'IH':
                present_total += 1

        return days, statuses, absent_total, present_total

class ExportMonthlyPayrollExcel(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        row_num = 1
        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                self.add_employee_data(ws, details, records, first_day_of_month, last_day_of_month, num_days, row_num)
                row_num += 4

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Payroll_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def add_employee_data(self, ws, details, records, first_day, last_day, num_days, start_row):
        """
        Add employee details and attendance data to the worksheet.
        """
        # Add employee header
        header = (
            f"Monthly Payroll Register for the Period of: {first_day} to {last_day}"
        )
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_days)
        cell = ws.cell(row=start_row, column=1, value=header)
        cell.font = Font(bold=True)

        # Process attendance records
        days, statuses, present_total, absent_total, ww_count, wo_count, total_working, total_late_entry, total_early_exit, total_overtime = self.process_records(records, first_day, num_days)

        # Define header labels
        header_labels = ["EMP ID", "EMP Name", "Department", "Designation", "Type", "Days"]
        header_labels_2 = ["Calender Days", "Working Days", "Paid", "WO", "WW", "PH", "FS", "CL", "EL", "SL", "Total Working", "Total Late Entry", "Total Early Exit", "Total Overtime"]
        row_data = header_labels + days + ["Present", "Absent"] + header_labels_2

        # Adding the rows to the sheet
        rows = [
            row_data,
            [f"{details['employee_id']}", f"{details['employee_name']}", f"{details['department']}", f"{details['designation']}", f"{details['job_type']}", "Status"] + statuses + 
            [present_total, absent_total, 0, 0, 0, wo_count, ww_count, 0, 0, 0, 0, 0, total_working, total_late_entry, total_early_exit, total_overtime],
        ]

        # Calculate max width for each column
        max_widths = []
        for col in range(len(header_labels)):
            max_width = max(len(str(header_labels[col])), len(str(rows[1][col])))
            max_widths.append(max_width)

        # Iterate through each row of data and add it to the worksheet
        for i, row_data in enumerate(rows, start=start_row + 1):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.THIN_BORDER
                
                # Style the header column (first columns)
                if col < 7:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    # Adjust column width based on max content width
                    if col <= len(max_widths):
                        ws.column_dimensions[cell.column_letter].width = max_widths[col-1] + 2  # Add padding

                    # Style the "Present" column (last but one)
                    if col == len(header_labels) + num_days + 1:  # This is the column for "Present"
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    
                    # Style the "Absent" column (last column)
                    if col == len(header_labels) + num_days + 2:  # This is the column for "Absent"
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

                    if col > len(header_labels) + num_days + 2:
                        max_width = max(len(str(header_labels_2[col])), len(str(rows[1][col])))
                        ws.column_dimensions[cell.column_letter].width = max_width + 2  
                
                # Style the status cells
                if i == start_row + 2:  # Status row
                    style = self.SHIFT_STATUS_STYLES.get(value)
                    if style:
                        cell.style = style
                
                # Style the days row
                if i == start_row + 1 and col > 6: 
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        statuses = []
        total_hours = []
        absent_total = 0
        present_total = 0
        ww_count = 0
        wo_count = 0

        total_working = timedelta(0)
        total_late_entry = timedelta(0)
        total_early_exit = timedelta(0)
        total_overtime = timedelta(0)

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                statuses.append(record.shift_status or "")
            else:
                statuses.append("")

        for status_code in statuses:
            if status_code == 'A':
                absent_total += 1
            elif status_code == 'HD':
                present_total += 0.5
                absent_total += 0.5
            elif status_code == 'P':
                present_total += 1
            elif status_code == 'WW':
                present_total += 1
            elif status_code == 'IH':
                present_total += 1
        
        # Count number of statuses WW and WO
        ww_count = statuses.count('WW')
        wo_count = statuses.count('WO') + statuses.count('WW')

        # Calculate totals safely
        total_working = sum((record.total_time or timedelta(0) for record in records), timedelta(0))
        total_late_entry = sum((record.late_entry or timedelta(0) for record in records), timedelta(0))
        total_early_exit = sum((record.early_exit or timedelta(0) for record in records), timedelta(0))
        total_overtime = sum((record.overtime or timedelta(0) for record in records), timedelta(0))

        # Format the output as HH:MM:SS
        total_working = str(total_working)
        total_late_entry = str(total_late_entry)
        total_early_exit = str(total_early_exit)
        total_overtime = str(total_overtime)

        return days, statuses, absent_total, present_total, ww_count, wo_count, total_working, total_late_entry, total_early_exit, total_overtime
    
class ExportMonthlyPayrollExcel2(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Payroll_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        # Add main header
        main_header = f"Monthly Payroll Register for the Period of: {first_day_of_month} to {last_day_of_month} | Printed Date: {datetime.now().strftime('%d-%m-%Y')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 6)
        cell = ws.cell(row=1, column=1, value=main_header)
        cell.font = Font(bold=True)

        # Add column headers
        header_labels = ["EMP ID", "EMP Name", "Company", "Location", "Department", "Designation", "Type", "Days"]
        days_header = [str(day) for day in range(1, num_days + 1)]
        total_header = ["Calender Days", "Working Days", "Paid Days", "Present", "Absent", "MP", "WW", "WO", "PH", "FS", "CL", "EL", "SL", "Total Working", "Total Late Entry", "Total Early Exit", "Total Overtime"]
        full_header = header_labels + days_header + total_header
        ws.append(full_header)

        # Style header row
        for col in range(1, len(full_header) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            cell.border = self.THIN_BORDER

            if col >= 9:  # Style days and total columns
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # Add employee data rows
        row_num = 3  # Start from row 3 after headers
        total_present = 0
        total_absent = 0
        ww_count = 0
        wo_count = 0
        grand_total_working_seconds = 0
        grand_total_late_entry = 0
        grand_total_early_exit = 0
        grand_total_overtime = 0

        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                days, statuses,total_days, working_days, paid_days, present_total, absent_total, mp_count, ww_count, wo_count, total_working, total_late_entry, total_early_exit, total_overtime = self.process_records(records, first_day_of_month, num_days)
                data_row = [
                    details['employee_id'],
                    details['employee_name'],
                    details['company'],
                    details['location'],
                    details['department'],
                    details['designation'],
                    details['job_type'],
                    "Status",
                ] + statuses + [total_days, working_days, paid_days, present_total, absent_total, mp_count, ww_count, wo_count, 0, 0, 0, 0, 0, total_working, total_late_entry, total_early_exit, total_overtime]
                ws.append(data_row)

                # Apply styling to data row
                for col in range(1, len(data_row) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = self.THIN_BORDER
                    if col <= 8:  # Style employee info columns
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    if col >= 7:  # Style days and total columns
                        cell.font = Font(bold=True)
                        if col < len(data_row) - 16:  # Style shift cells
                            shift = data_row[col - 1]
                            style = self.SHIFT_STATUS_STYLES.get(shift)
                            if style:
                                cell.style = style
                        else:  # Style total columns
                            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

                row_num += 1

                total_present += present_total
                total_absent += absent_total
                ww_count += ww_count
                wo_count += wo_count

                # Accumulate grand total seconds
                if total_working and total_working != 'None':
                    hours, minutes, seconds = map(int, total_working.split(':'))
                    grand_total_working_seconds += hours * 3600 + minutes * 60 + seconds
                
                # Check and handle None for time values
                if total_late_entry and total_late_entry != 'None':
                    hours, minutes, seconds = map(int, total_late_entry.split(':'))
                    grand_total_late_entry += hours * 3600 + minutes * 60 + seconds
                
                if total_early_exit and total_early_exit != 'None':
                    hours, minutes, seconds = map(int, total_early_exit.split(':'))
                    grand_total_early_exit += hours * 3600 + minutes * 60 + seconds
                
                if total_overtime and total_overtime != 'None':
                    hours, minutes, seconds = map(int, total_overtime.split(':'))
                    grand_total_overtime += hours * 3600 + minutes * 60 + seconds
        
        # Calculate grand total hours
        grand_total_working_hours = self.format_timedelta_to_hhmmss(grand_total_working_seconds)
        grand_total_late_entry = self.format_timedelta_to_hhmmss(grand_total_late_entry)
        grand_total_early_exit = self.format_timedelta_to_hhmmss(grand_total_early_exit)
        grand_total_overtime = self.format_timedelta_to_hhmmss(grand_total_overtime)

        # Add grand total row
        grand_total_row = [""] * (len(full_header) - 17) + ["Grand Total"] + [total_present, total_absent, ww_count, wo_count, 0, 0, 0, 0, 0, grand_total_working_hours, grand_total_late_entry, grand_total_early_exit, grand_total_overtime]
        ws.append(grand_total_row)

        # Style grand total row
        for col in range(len(full_header) - 17, len(full_header) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color
            cell.border = self.THIN_BORDER

        # Adjust column widths
        self.adjust_column_widths(ws)

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Payroll_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def format_timedelta_to_hhmmss(self, td):
        """
        Convert a timedelta to HH:MM:SS format, handling days correctly.
        
        Args:
            td (timedelta): The timedelta to convert
            
        Returns:
            str: Time in HH:MM:SS format
        """
        if not td:
            return None
        
        if isinstance(td, int):
          total_seconds = td
        else:
            # Convert timedelta to total seconds
            total_seconds = int(td.total_seconds())

        # If total seconds is 0, return None
        if total_seconds == timedelta(0):
            return None
        
        # Calculate hours including days
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        total_days = len(days)
        statuses = []
        present_total = 0
        absent_total = 0
        mp_count = 0 
        ww_count = 0
        wo_count = 0

        total_working = timedelta(0)
        total_late_entry = timedelta(0)
        total_early_exit = timedelta(0)
        total_overtime = timedelta(0)

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                statuses.append(record.shift_status or "")
            else:
                statuses.append("")

        for status_code in statuses:
            if status_code == 'A':
                absent_total += 1
            elif status_code == 'HD':
                present_total += 0.5
                absent_total += 0.5
            elif status_code == 'P':
                present_total += 1
            elif status_code == 'IH':
                present_total += 1
        
        # Count number of statuses WW and WO and MP
        ww_count = statuses.count('WW')
        wo_count = statuses.count('WO') + statuses.count('WW')
        mp_count = statuses.count('MP')
        
        # Calculate totals safely
        total_working_le = sum((record.total_time or timedelta(0) for record in records), timedelta(0))
        total_late_entry_le = sum((record.late_entry or timedelta(0) for record in records), timedelta(0))
        total_early_exit_le = sum((record.early_exit or timedelta(0) for record in records), timedelta(0))
        total_overtime_le = sum((record.overtime or timedelta(0) for record in records), timedelta(0))

        total_working = self.format_timedelta_to_hhmmss(total_working_le)
        total_late_entry = self.format_timedelta_to_hhmmss(total_late_entry_le)
        total_early_exit = self.format_timedelta_to_hhmmss(total_early_exit_le)
        total_overtime = self.format_timedelta_to_hhmmss(total_overtime_le)

        # Format the output as HH:MM:SS
        total_working = str(total_working) if total_working else None
        total_late_entry = str(total_late_entry) if total_late_entry else None
        total_early_exit = str(total_early_exit) if total_early_exit else None
        total_overtime = str(total_overtime) if total_overtime else None

        working_days = total_days - wo_count
        # paid_days = working_days - absent_total - mp_count
        paid_days = present_total + wo_count

        return days, statuses, total_days, working_days, paid_days, present_total, absent_total, mp_count, ww_count, wo_count, total_working, total_late_entry, total_early_exit, total_overtime

    def adjust_column_widths(self, ws):
        """
        Adjust column widths based on the maximum content length in each column,
        skipping merged cells and applying a minimum and maximum width limit.
        """
        max_width = {}
        max_allowed_width = 30  # Set a maximum width limit for columns
        min_allowed_width = 10  # Set a minimum width limit for columns

        for row in ws.iter_rows():
            for cell in row:
                # Skip merged cells as they don't have column_letter
                if isinstance(cell, MergedCell):
                    continue
                column = cell.column_letter
                value_length = len(str(cell.value)) if cell.value is not None else 0
                if column not in max_width or value_length > max_width[column]:
                    max_width[column] = min(value_length, max_allowed_width)  # Apply max width limit

        for col_letter, width in max_width.items():
            # Apply minimum and maximum width, adding padding
            adjusted_width = max(min_allowed_width, width + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

class ExportMonthlyShiftRoasterExcel(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        row_num = 1
        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                self.add_employee_data(ws, details, records, first_day_of_month, last_day_of_month, num_days, row_num)
                row_num += 4

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Shift_Roaster_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def add_employee_data(self, ws, details, records, first_day, last_day, num_days, start_row):
        """
        Add employee details and attendance data to the worksheet.
        """
        # Add employee header
        header = (
            f"Monthly Shift Roaster Register for the Period of: {first_day} to {last_day}"
        )
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_days)
        cell = ws.cell(row=start_row, column=1, value=header)
        cell.font = Font(bold=True)

        # Process attendance records
        days, shifts = self.process_records(records, first_day, num_days)

        # Define header labels
        header_labels = ["EMP ID", "EMP Name", "Department", "Designation", "Type", "Days"]
        row_data = header_labels + days + ["FS", "SS", "NS", "GS"]

        # Adding the rows to the sheet
        rows = [
            row_data,
            [f"{details['employee_id']}", f"{details['employee_name']}", f"{details['department']}", f"{details['designation']}", f"{details['job_type']}", "Shift"] + shifts + 
            [0, 0, 0, 0],
        ]

        # Calculate max width for each column
        max_widths = []
        for col in range(len(header_labels)):
            max_width = max(len(str(header_labels[col])), len(str(rows[1][col])))
            max_widths.append(max_width)

        # Iterate through each row of data and add it to the worksheet
        for i, row_data in enumerate(rows, start=start_row + 1):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.THIN_BORDER
                
                # Style the header column (first columns)
                if col < 7:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    # Adjust column width based on max content width
                    if col <= len(max_widths):
                        ws.column_dimensions[cell.column_letter].width = max_widths[col-1] + 2  # Add padding
                
                # Style the status cells
                if i == start_row + 2:  # Status row
                    cell.font = Font(bold=True)
                
                # Style the days row
                if i == start_row + 1 and col > 6: 
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        shifts = []

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                shifts.append(record.shift or "")
            else:
                shifts.append("")

        return days, shifts
    
class ExportMonthlyShiftRoasterExcel2(View):
    """
    View to generate and export an Excel file containing monthly shift roaster details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Shift_Roaster_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        # Add main header
        main_header = f"Monthly Shift Roaster Register for the Period of: {first_day_of_month} to {last_day_of_month} | Printed Date: {datetime.now().strftime('%d-%m-%Y')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 6)
        cell = ws.cell(row=1, column=1, value=main_header)
        cell.font = Font(bold=True)

        # Add column headers
        header_labels = ["EMP ID", "EMP Name", "Company", "Location", "Department", "Designation", "Type", "Days"]
        days_header = [str(day) for day in range(1, num_days + 1)]
        total_header = ["FS", "SS", "NS", "GS"]
        full_header = header_labels + days_header + total_header
        ws.append(full_header)

        # Style header row
        for col in range(1, len(full_header) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            cell.border = self.THIN_BORDER

            if col >= 9:  # Style days and total columns
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # Add employee data rows
        row_num = 3  # Start from row 3 after headers
        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                days, shifts, fs_count, ss_count, ns_count, gs_count = self.process_records(records, first_day_of_month, num_days)
                data_row = [
                    details['employee_id'],
                    details['employee_name'],
                    details['company'],
                    details['location'],
                    details['department'],
                    details['designation'],
                    details['job_type'],
                    "Shift",
                ] + shifts + [fs_count, ss_count, ns_count, gs_count]  # Placeholder for FS, SS, NS, GS
                ws.append(data_row)

                # Apply styling to data row
                for col in range(1, len(data_row) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = self.THIN_BORDER
                    if col <= 8:  # Style employee info columns
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    if col >= 7:  # Style days and total columns
                        cell.font = Font(bold=True)
                        if col < len(data_row) - 3:  # Style shift cells
                            shift = data_row[col - 1]
                            style = self.SHIFT_STATUS_STYLES.get(shift)
                            if style:
                                cell.style = style
                        else:  # Style FS, SS, NS, GS columns
                            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

                row_num += 1

        # Adjust column widths
        self.adjust_column_widths(ws)

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Shift_Roaster_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        shifts = []

        fs_count = 0
        ss_count = 0
        ns_count = 0
        gs_count = 0

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                shift = record.shift or ""
                shifts.append(shift)
                if shift == "FS":  # Check if the shift is 'FS'
                    fs_count += 1
                elif shift == "SS":
                    ss_count += 1
                elif shift == "NS":
                    ns_count += 1   
                elif shift == "GS":
                    gs_count += 1
            else:
                shifts.append("")

        return days, shifts, fs_count, ss_count, ns_count, gs_count

    def adjust_column_widths(self, ws):
        """
        Adjust column widths based on the maximum content length in each column,
        skipping merged cells and applying a minimum and maximum width limit.
        """
        max_width = {}
        max_allowed_width = 30  # Set a maximum width limit for columns
        min_allowed_width = 10  # Set a minimum width limit for columns

        for row in ws.iter_rows():
            for cell in row:
                # Skip merged cells as they don't have column_letter
                if isinstance(cell, MergedCell):
                    continue
                column = cell.column_letter
                value_length = len(str(cell.value)) if cell.value is not None else 0
                if column not in max_width or value_length > max_width[column]:
                    max_width[column] = min(value_length, max_allowed_width)  # Apply max width limit

        for col_letter, width in max_width.items():
            # Apply minimum and maximum width, adding padding
            adjusted_width = max(min_allowed_width, width + 2)
            ws.column_dimensions[col_letter].width = adjusted_width
    
class ExportMonthlyOvertimeExcel(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        row_num = 1
        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                self.add_employee_data(ws, details, records, first_day_of_month, last_day_of_month, num_days, row_num)
                row_num += 4

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Overtime_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def add_employee_data(self, ws, details, records, first_day, last_day, num_days, start_row):
        """
        Add employee details and attendance data to the worksheet.
        """
        # Add employee header
        header = (
            f"Monthly Overtime Register for the Period of: {first_day} to {last_day}"
        )
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_days)
        cell = ws.cell(row=start_row, column=1, value=header)
        cell.font = Font(bold=True)

        # Process attendance records
        days, overtime, formatted_total_overtime = self.process_records(records, first_day, num_days)

        # Define header labels
        header_labels = ["EMP ID", "EMP Name", "Department", "Designation", "Type", "Days"]
        row_data = header_labels + days + ["Total"]

        # Adding the rows to the sheet
        rows = [
            row_data,
            [f"{details['employee_id']}", f"{details['employee_name']}", f"{details['department']}", f"{details['designation']}", f"{details['job_type']}", "Overtime"] + overtime + 
            [formatted_total_overtime],
        ]

        # Calculate max width for each column
        max_widths = []
        for col in range(len(header_labels)):
            max_width = max(len(str(header_labels[col])), len(str(rows[1][col])))
            max_widths.append(max_width)

        # Iterate through each row of data and add it to the worksheet
        for i, row_data in enumerate(rows, start=start_row + 1):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.THIN_BORDER
                
                # Style the header column (first columns)
                if col < 7:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    # Adjust column width based on max content width
                    if col <= len(max_widths):
                        ws.column_dimensions[cell.column_letter].width = max_widths[col-1] + 2  # Add padding
                
                # Style the status cells
                if i == start_row + 2:  # Status row
                    cell.font = Font(bold=True)
                
                # Style the days row
                if i == start_row + 1 and col > 6: 
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        overtime = []
        total_overtime = timedelta(0)

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                total_overtime_value = record.overtime or timedelta(0)
                overtime.append(str(total_overtime_value if total_overtime_value != timedelta(0) else ""))
                total_overtime += total_overtime_value
            else:
                overtime.append("")

        # Format total overtime time as hours, minutes, and seconds
        if total_overtime == timedelta(0):
            formatted_total_overtime = "0"
        else:
            formatted_total_overtime = (
                f"{total_overtime.seconds // 3600}:"
                f"{(total_overtime.seconds % 3600) // 60}:"
                f"{total_overtime.seconds % 60}"
            )

        return days, overtime, formatted_total_overtime
    
class ExportMonthlyOvertimeExcel2(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day


        # Add main header
        main_header = f"Monthly Overtime Register for the Period of: {first_day_of_month} to {last_day_of_month} | Printed Date: {datetime.now().strftime('%d-%m-%Y')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 9)
        cell = ws.cell(row=1, column=1, value=main_header)
        cell.font = Font(bold=True)

        # Add column headers
        header_labels = ["EMP ID", "EMP Name", "Company", "Location", "Department", "Designation", "Type", "Days"]
        days_header = [str(day) for day in range(1, num_days + 1)]
        total_header = ["Total"]
        full_header = header_labels + days_header + total_header
        ws.append(full_header)

        # Style header row
        for col in range(1, len(full_header) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            cell.border = self.THIN_BORDER
        
            if col >= 9:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                
                # Adjust column width based on max content width
                # if col <= len(max_widths):
                #     ws.column_dimensions[cell.column_letter].width = max_widths[col-1] + 2
            
        # Add employee data rows
        row_num = 3  # Start from row 3 after headers
        grand_total_seconds = 0

        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                overtime, formatted_total = self.process_records(records, first_day_of_month, num_days)
                data_row = [
                    details['employee_id'],
                    details['employee_name'],
                    details['company'],
                    details['location'],
                    details['department'],
                    details['designation'],
                    details['job_type'],
                    "Overtime",
                ] + overtime + [formatted_total]
                ws.append(data_row)

                # Apply styling to data row
                for col in range(1, len(data_row) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = self.THIN_BORDER
                    if col <= 8:  # Style employee info columns
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    if col >= 9:  # Style days and total columns
                        cell.font = Font(bold=True)
                
                row_num += 1

                # Accumulate grand total seconds
                if formatted_total:
                    hours, minutes, seconds = map(int, formatted_total.split(':'))
                    grand_total_seconds += hours * 3600 + minutes * 60 + seconds
        
        # Calculate grand total hours
        grand_total_hours = grand_total_seconds // 3600
        grand_total_minutes = (grand_total_seconds % 3600) // 60
        grand_total_seconds = grand_total_seconds % 60
        grand_total_str = f"{grand_total_hours}:{grand_total_minutes:02d}:{grand_total_seconds:02d}"

        # Add grand total row
        grand_total_row = [""] * (len(full_header) - 2) + ["Grand Total"] + [grand_total_str]
        ws.append(grand_total_row)

        # Style grand total row
        for col in range(len(full_header) - 1, len(full_header) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color
            cell.border = self.THIN_BORDER

        # Adjust column widths
        self.adjust_column_widths(ws)

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Overtime_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records to extract overtime for each day and total.
        Returns overtime for each day and the total overtime in "HH:MM:SS" format,
        even if it exceeds 24 hours (e.g., converts "1 day 03:39:42" to "27:39:42").
        """
        overtime = []
        total_overtime = timedelta(0)
        grand_total_seconds = 0

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            
            if record and record.overtime:
                # Accumulate total overtime
                total_overtime += record.overtime

                # Convert individual day's overtime to "HH:MM:SS"
                ot_seconds = int(record.overtime.total_seconds())
                hours = ot_seconds // 3600  # Total hours (including days)
                minutes = (ot_seconds % 3600) // 60
                seconds = ot_seconds % 60
                ot_str = f"{hours}:{minutes:02d}:{seconds:02d}"
                overtime.append(ot_str)
            else:
                overtime.append("")  # No overtime for this day

        # Format total overtime into "HH:MM:SS"
        if total_overtime == timedelta(0):
            formatted_total = None  # No overtime
        else:
            total_seconds = int(total_overtime.total_seconds())
            hours = total_seconds // 3600  # Total hours (including days)
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            formatted_total = f"{hours}:{minutes:02d}:{seconds:02d}"

        return overtime, formatted_total

    def adjust_column_widths(self, ws):
        """
        Adjust column widths based on the maximum content length in each column,
        skipping merged cells and applying a minimum and maximum width limit.
        """
        max_width = {}
        max_allowed_width = 30  # Set a maximum width limit for columns
        min_allowed_width = 10  # Set a minimum width limit for columns

        for row in ws.iter_rows():
            for cell in row:
                # Skip merged cells as they don't have column_letter
                if isinstance(cell, MergedCell):
                    continue
                column = cell.column_letter
                value_length = len(str(cell.value)) if cell.value is not None else 0
                if column not in max_width or value_length > max_width[column]:
                    max_width[column] = min(value_length, max_allowed_width)  # Apply max width limit

        for col_letter, width in max_width.items():
            # Apply minimum and maximum width, adding padding
            adjusted_width = max(min_allowed_width, width + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

class ExportMonthlyOvertimeRoundoffExcel2(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day


        # Add main header
        main_header = f"Monthly Overtime Roundoff Register for the Period of: {first_day_of_month} to {last_day_of_month} | Printed Date: {datetime.now().strftime('%d-%m-%Y')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 9)
        cell = ws.cell(row=1, column=1, value=main_header)
        cell.font = Font(bold=True)

        # Add column headers
        header_labels = ["EMP ID", "EMP Name", "Company", "Location", "Department", "Designation", "Type", "Days"]
        days_header = [str(day) for day in range(1, num_days + 1)]
        total_header = ["Total"]
        full_header = header_labels + days_header + total_header
        ws.append(full_header)

        # Style header row
        for col in range(1, len(full_header) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            cell.border = self.THIN_BORDER
        
            if col >= 9:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                
                # Adjust column width based on max content width
                # if col <= len(max_widths):
                #     ws.column_dimensions[cell.column_letter].width = max_widths[col-1] + 2
            
        # Add employee data rows
        row_num = 3  # Start from row 3 after headers
        grand_total_seconds = 0

        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                overtime, formatted_total = self.process_records(records, first_day_of_month, num_days)
                data_row = [
                    details['employee_id'],
                    details['employee_name'],
                    details['company'],
                    details['location'],
                    details['department'],
                    details['designation'],
                    details['job_type'],
                    "Overtime",
                ] + overtime + [formatted_total]
                ws.append(data_row)

                # Apply styling to data row
                for col in range(1, len(data_row) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = self.THIN_BORDER
                    if col <= 8:  # Style employee info columns
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    if col >= 9:  # Style days and total columns
                        cell.font = Font(bold=True)
                
                row_num += 1

                # Accumulate grand total seconds
                if formatted_total:
                    hours, minutes, seconds = map(int, formatted_total.split(':'))
                    grand_total_seconds += hours * 3600 + minutes * 60 + seconds
        
        # Calculate grand total hours
        grand_total_hours = grand_total_seconds // 3600
        grand_total_minutes = (grand_total_seconds % 3600) // 60
        grand_total_seconds = grand_total_seconds % 60
        grand_total_str = f"{grand_total_hours}:{grand_total_minutes:02d}:{grand_total_seconds:02d}"

        # Add grand total row
        grand_total_row = [""] * (len(full_header) - 2) + ["Grand Total"] + [grand_total_str]
        ws.append(grand_total_row)

        # Style grand total row
        for col in range(len(full_header) - 1, len(full_header) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color
            cell.border = self.THIN_BORDER

        # Adjust column widths
        self.adjust_column_widths(ws)

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Roundoff_Overtime_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records to extract overtime for each day and total.
        Returns overtime for each day and the total overtime in "HH:MM:SS" format,
        even if it exceeds 24 hours (e.g., converts "1 day 03:39:42" to "27:39:42").
        """
        overtime = []
        total_overtime = timedelta(0)
        grand_total_seconds = 0

        # Get the singleton instance of OvertimeRoundoffRules
        overtime_rules = OvertimeRoundoffRules.load()

        # Access the attributes from the instance:
        rounding_interval = overtime_rules.round_off_interval
        rounding_direction = overtime_rules.round_off_direction
        rounding_threshold = rounding_interval / 2 

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            
            if record and record.overtime:
                # Convert overtime to total seconds
                ot_seconds = int(record.overtime.total_seconds())

                # Use the instance attributes for calculations:
                remainder = ot_seconds % rounding_interval.total_seconds()

                if rounding_direction.lower() == "nearest":
                    if remainder >= rounding_threshold.total_seconds():
                        rounded_seconds = int(ot_seconds + (rounding_interval.total_seconds() - remainder))
                    else:
                        rounded_seconds = int(ot_seconds - remainder)

                elif rounding_direction.lower() == "up":
                    rounded_seconds = int(ot_seconds + (rounding_interval.total_seconds() - remainder)) if remainder > 0 else ot_seconds

                else:
                    rounded_seconds = int(ot_seconds - remainder)

                # Convert rounded seconds back to timedelta
                rounded_overtime = timedelta(seconds=rounded_seconds)

                # Accumulate total overtime
                total_overtime += rounded_overtime

                # Convert individual day's overtime to "HH:MM:SS"
                hours = rounded_seconds // 3600  # Total hours (including days)
                minutes = (rounded_seconds % 3600) // 60
                seconds = rounded_seconds % 60
                ot_str = f"{hours}:{minutes:02d}:{seconds:02d}"
                overtime.append(ot_str)
            else:
                overtime.append("")  # No overtime for this day

        # Format total overtime into "HH:MM:SS"
        if total_overtime == timedelta(0):
            formatted_total = None  # No overtime
        else:
            total_seconds = int(total_overtime.total_seconds())
            hours = total_seconds // 3600  # Total hours (including days)
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            formatted_total = f"{hours}:{minutes:02d}:{seconds:02d}"

        return overtime, formatted_total

    def adjust_column_widths(self, ws):
        """
        Adjust column widths based on the maximum content length in each column,
        skipping merged cells and applying a minimum and maximum width limit.
        """
        max_width = {}
        max_allowed_width = 30  # Set a maximum width limit for columns
        min_allowed_width = 10  # Set a minimum width limit for columns

        for row in ws.iter_rows():
            for cell in row:
                # Skip merged cells as they don't have column_letter
                if isinstance(cell, MergedCell):
                    continue
                column = cell.column_letter
                value_length = len(str(cell.value)) if cell.value is not None else 0
                if column not in max_width or value_length > max_width[column]:
                    max_width[column] = min(value_length, max_allowed_width)  # Apply max width limit

        for col_letter, width in max_width.items():
            # Apply minimum and maximum width, adding padding
            adjusted_width = max(min_allowed_width, width + 2)
            ws.column_dimensions[col_letter].width = adjusted_width
    
class ExportMonthlyLateEntryExcel(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        row_num = 1
        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                self.add_employee_data(ws, details, records, first_day_of_month, last_day_of_month, num_days, row_num)
                row_num += 4

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Late_Entry_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def add_employee_data(self, ws, details, records, first_day, last_day, num_days, start_row):
        """
        Add employee details and attendance data to the worksheet.
        """
        # Add employee header
        header = (
            f"Monthly Late Entry Register for the Period of: {first_day} to {last_day}"
        )
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_days)
        cell = ws.cell(row=start_row, column=1, value=header)
        cell.font = Font(bold=True)

        # Process attendance records
        days, late_entry, formatted_total_late_entry = self.process_records(records, first_day, num_days)

        # Define header labels
        header_labels = ["EMP ID", "EMP Name", "Department", "Designation", "Type", "Days"]
        row_data = header_labels + days + ["Total"]

        # Adding the rows to the sheet
        rows = [
            row_data,
            [f"{details['employee_id']}", f"{details['employee_name']}", f"{details['department']}", f"{details['designation']}", f"{details['job_type']}", "Late Entry Hrs"] + late_entry + 
            [formatted_total_late_entry],
        ]

        # Calculate max width for each column
        max_widths = []
        for col in range(len(header_labels)):
            max_width = max(len(str(header_labels[col])), len(str(rows[1][col])))
            max_widths.append(max_width)

        # Iterate through each row of data and add it to the worksheet
        for i, row_data in enumerate(rows, start=start_row + 1):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.THIN_BORDER
                
                # Style the header column (first columns)
                if col < 7:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    # Adjust column width based on max content width
                    if col <= len(max_widths):
                        ws.column_dimensions[cell.column_letter].width = max_widths[col-1] + 2  # Add padding
                
                # Style the status cells
                if i == start_row + 2:  # Status row
                    cell.font = Font(bold=True)
                
                # Style the days row
                if i == start_row + 1 and col > 6: 
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        late_entry = []
        total_late_entry = timedelta(0)

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                late_entry_value = record.late_entry or timedelta(0)
                late_entry.append(str(late_entry_value if late_entry_value != timedelta(0) else ""))
                total_late_entry += late_entry_value
            else:
                late_entry.append("")

        # Format total late entry time as hours, minutes, and seconds
        if total_late_entry == timedelta(0):
            formatted_total_late_entry = "0"
        else:
            formatted_total_late_entry = (
                f"{total_late_entry.seconds // 3600}:"
                f"{(total_late_entry.seconds % 3600) // 60}:"
                f"{total_late_entry.seconds % 60}"
            )

        return days, late_entry, formatted_total_late_entry
    
class ExportMonthlyLateEntryExcel2(View):
    """
    View to generate and export an Excel file containing monthly late entry details for all employees.
    """

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Late_Entry_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        # Add main header
        main_header = f"Monthly Late Entry Register for the Period of: {first_day_of_month} to {last_day_of_month} | Printed Date: {datetime.now().strftime('%d-%m-%Y')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 8)
        cell = ws.cell(row=1, column=1, value=main_header)
        cell.font = Font(bold=True)

        # Add column headers
        header_labels = ["EMP ID", "EMP Name", "Company", "Location", "Department", "Designation", "Type", "Days"]
        days_header = [str(day) for day in range(1, num_days + 1)]
        total_header = ["Total"]
        full_header = header_labels + days_header + total_header
        ws.append(full_header)

        # Style header row
        for col in range(1, len(full_header) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            cell.border = self.THIN_BORDER

            if col >= 9:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # Add employee data rows
        row_num = 3  # Start from row 3 after headers
        grand_total_seconds = 0

        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                late_entry, formatted_total = self.process_records(records, first_day_of_month, num_days)
                data_row = [
                    details['employee_id'],
                    details['employee_name'],
                    details['company'],
                    details['location'],
                    details['department'],
                    details['designation'],
                    details['job_type'],
                    "Late Entry",
                ] + late_entry + [formatted_total]
                ws.append(data_row)

                # Apply styling to data row
                for col in range(1, len(data_row) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = self.THIN_BORDER
                    if col <= 8:  # Style employee info columns
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    if col >= 9:
                        cell.font = Font(bold=True)

                row_num += 1

                # Accumulate grand total seconds
                if formatted_total:
                    hours, minutes, seconds = map(int, formatted_total.split(':'))
                    grand_total_seconds += hours * 3600 + minutes * 60 + seconds
        
        # Calculate grand total hours
        grand_total_hours = grand_total_seconds // 3600
        grand_total_minutes = (grand_total_seconds % 3600) // 60
        grand_total_seconds = grand_total_seconds % 60
        grand_total_str = f"{grand_total_hours}:{grand_total_minutes:02d}:{grand_total_seconds:02d}"

        # Add grand total row
        grand_total_row = [""] * (len(full_header) - 2) + ["Grand Total"] + [grand_total_str]
        ws.append(grand_total_row)

        # Style grand total row
        for col in range(len(full_header) - 1, len(full_header) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color
            cell.border = self.THIN_BORDER

        # Adjust column widths
        self.adjust_column_widths(ws)

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Late_Entry_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records to extract late entry for each day and total.
        Returns late entry for each day and the total in "HH:MM:SS" format.
        """
        late_entry = []
        total_late = timedelta(0)

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            
            if record and record.late_entry:
                total_late += record.late_entry
                le_seconds = int(record.late_entry.total_seconds())
                hours = le_seconds // 3600
                minutes = (le_seconds % 3600) // 60
                seconds = le_seconds % 60
                late_entry.append(f"{hours}:{minutes:02d}:{seconds:02d}")
            else:
                late_entry.append("")

        # Format total late entry
        if total_late == timedelta(0):
            formatted_total = None
        else:
            total_seconds = int(total_late.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            formatted_total = f"{hours}:{minutes:02d}:{seconds:02d}"

        return late_entry, formatted_total
    
    def adjust_column_widths(self, ws):
        """
        Adjust column widths based on the maximum content length in each column,
        skipping merged cells and applying a minimum and maximum width limit.
        """
        max_width = {}
        max_allowed_width = 30  # Set a maximum width limit for columns
        min_allowed_width = 10  # Set a minimum width limit for columns

        for row in ws.iter_rows():
            for cell in row:
                # Skip merged cells as they don't have column_letter
                if isinstance(cell, MergedCell):
                    continue
                column = cell.column_letter
                value_length = len(str(cell.value)) if cell.value is not None else 0
                if column not in max_width or value_length > max_width[column]:
                    max_width[column] = min(value_length, max_allowed_width)  # Apply max width limit

        for col_letter, width in max_width.items():
            # Apply minimum and maximum width, adding padding
            adjusted_width = max(min_allowed_width, width + 2)
            ws.column_dimensions[col_letter].width = adjusted_width
    
class ExportMonthlyEarlyExitExcel(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        row_num = 1
        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                self.add_employee_data(ws, details, records, first_day_of_month, last_day_of_month, num_days, row_num)
                row_num += 4

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Early_Exit_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def add_employee_data(self, ws, details, records, first_day, last_day, num_days, start_row):
        """
        Add employee details and attendance data to the worksheet.
        """
        # Add employee header
        header = (
            f"Monthly Early Exit Register for the Period of: {first_day} to {last_day}"
        )
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_days)
        cell = ws.cell(row=start_row, column=1, value=header)
        cell.font = Font(bold=True)

        # Process attendance records
        days, early_exit, formatted_total_early_exit = self.process_records(records, first_day, num_days)

        # Define header labels
        header_labels = ["EMP ID", "EMP Name", "Department", "Designation", "Type", "Days"]
        row_data = header_labels + days + ["Total"]

        # Adding the rows to the sheet
        rows = [
            row_data,
            [f"{details['employee_id']}", f"{details['employee_name']}", f"{details['department']}", f"{details['designation']}", f"{details['job_type']}", "Early Exit Hrs"] + early_exit + 
            [formatted_total_early_exit],
        ]

        # Calculate max width for each column
        max_widths = []
        for col in range(len(header_labels)):
            max_width = max(len(str(header_labels[col])), len(str(rows[1][col])))
            max_widths.append(max_width)

        # Iterate through each row of data and add it to the worksheet
        for i, row_data in enumerate(rows, start=start_row + 1):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.THIN_BORDER
                
                # Style the header column (first columns)
                if col < 7:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    # Adjust column width based on max content width
                    if col <= len(max_widths):
                        ws.column_dimensions[cell.column_letter].width = max_widths[col-1] + 2  # Add padding
                
                # Style the status cells
                if i == start_row + 2:  # Status row
                    cell.font = Font(bold=True)
                
                # Style the days row
                if i == start_row + 1 and col > 6: 
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        early_exit = []
        total_early_exit = timedelta(0)  # Initialize total as timedelta

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                early_exit_value = record.early_exit or timedelta(0)
                early_exit.append(str(early_exit_value if early_exit_value != timedelta(0) else ""))
                total_early_exit += early_exit_value
            else:
                early_exit.append("")

        # Format total early exit time as hours, minutes, and seconds
        if total_early_exit == timedelta(0):
            formatted_total_early_exit = "0"
        else:
            formatted_total_early_exit = (
                f"{total_early_exit.seconds // 3600}:"
                f"{(total_early_exit.seconds % 3600) // 60}:"
                f"{total_early_exit.seconds % 60}"
            )

        return days, early_exit, formatted_total_early_exit
    
class ExportMonthlyEarlyExitExcel2(View):
    """
    View to generate and export an Excel file containing monthly late entry details for all employees.
    """

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Early_Exit_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        # Add main header
        main_header = f"Monthly Early Exit Register for the Period of: {first_day_of_month} to {last_day_of_month} | Printed Date: {datetime.now().strftime('%d-%m-%Y')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 8)
        cell = ws.cell(row=1, column=1, value=main_header)
        cell.font = Font(bold=True)

        # Add column headers
        header_labels = ["EMP ID", "EMP Name", "Company", "Location", "Department", "Designation", "Type", "Days"]
        days_header = [str(day) for day in range(1, num_days + 1)]
        total_header = ["Total"]
        full_header = header_labels + days_header + total_header
        ws.append(full_header)

        # Style header row
        for col in range(1, len(full_header) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            cell.border = self.THIN_BORDER

            if col >= 9:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # Add employee data rows
        row_num = 3  # Start from row 3 after headers
        grand_total_seconds = 0

        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                early_exit, formatted_total = self.process_records(records, first_day_of_month, num_days)
                data_row = [
                    details['employee_id'],
                    details['employee_name'],
                    details['company'],
                    details['location'],
                    details['department'],
                    details['designation'],
                    details['job_type'],
                    "Early Exit",
                ] + early_exit + [formatted_total]
                ws.append(data_row)

                # Apply styling to data row
                for col in range(1, len(data_row) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = self.THIN_BORDER
                    if col <= 8:  # Style employee info columns
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    if col >= 9:
                        cell.font = Font(bold=True)

                row_num += 1

                # Accumulate grand total seconds
                if formatted_total:
                    hours, minutes, seconds = map(int, formatted_total.split(':'))
                    grand_total_seconds += hours * 3600 + minutes * 60 + seconds
        
        # Calculate grand total hours
        grand_total_hours = grand_total_seconds // 3600
        grand_total_minutes = (grand_total_seconds % 3600) // 60
        grand_total_seconds = grand_total_seconds % 60
        grand_total_str = f"{grand_total_hours}:{grand_total_minutes:02d}:{grand_total_seconds:02d}"

        # Add grand total row
        grand_total_row = [""] * (len(full_header) - 2) + ["Grand Total"] + [grand_total_str]
        ws.append(grand_total_row)

        # Style grand total row
        for col in range(len(full_header) - 1, len(full_header) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color
            cell.border = self.THIN_BORDER

        # Adjust column widths
        self.adjust_column_widths(ws)

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Early_Exit_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records to extract late entry for each day and total.
        Returns late entry for each day and the total in "HH:MM:SS" format.
        """
        early_exit = []
        total_early = timedelta(0)

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            
            if record and record.early_exit:
                total_early += record.early_exit
                le_seconds = int(record.early_exit.total_seconds())
                hours = le_seconds // 3600
                minutes = (le_seconds % 3600) // 60
                seconds = le_seconds % 60
                early_exit.append(f"{hours}:{minutes:02d}:{seconds:02d}")
            else:
                early_exit.append("")

        # Format total late entry
        if total_early == timedelta(0):
            formatted_total = None
        else:
            total_seconds = int(total_early.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            formatted_total = f"{hours}:{minutes:02d}:{seconds:02d}"

        return early_exit, formatted_total
    
    def adjust_column_widths(self, ws):
        """
        Adjust column widths based on the maximum content length in each column,
        skipping merged cells and applying a minimum and maximum width limit.
        """
        max_width = {}
        max_allowed_width = 30  # Set a maximum width limit for columns
        min_allowed_width = 10  # Set a minimum width limit for columns

        for row in ws.iter_rows():
            for cell in row:
                # Skip merged cells as they don't have column_letter
                if isinstance(cell, MergedCell):
                    continue
                column = cell.column_letter
                value_length = len(str(cell.value)) if cell.value is not None else 0
                if column not in max_width or value_length > max_width[column]:
                    max_width[column] = min(value_length, max_allowed_width)  # Apply max width limit

        for col_letter, width in max_width.items():
            # Apply minimum and maximum width, adding padding
            adjusted_width = max(min_allowed_width, width + 2)
            ws.column_dimensions[col_letter].width = adjusted_width
    
class ExportMonthlyAbsentExcel(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        row_num = 1
        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                self.add_employee_data(ws, details, records, first_day_of_month, last_day_of_month, num_days, row_num)
                row_num += 4

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Absent_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def add_employee_data(self, ws, details, records, first_day, last_day, num_days, start_row):
        """
        Add employee details and attendance data to the worksheet.
        """
        # Add employee header
        header = (
            f"Monthly Late Entry Register for the Period of: {first_day} to {last_day}"
        )
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_days)
        cell = ws.cell(row=start_row, column=1, value=header)
        cell.font = Font(bold=True)

        # Process attendance records
        days, statuses, absent_total = self.process_records(records, first_day, num_days)

        # Define header labels
        header_labels = ["EMP ID", "EMP Name", "Department", "Designation", "Type", "Days"]
        row_data = header_labels + days + ["Absent Days"]

        # Adding the rows to the sheet
        rows = [
            row_data,
            [f"{details['employee_id']}", f"{details['employee_name']}", f"{details['department']}", f"{details['designation']}", f"{details['job_type']}", "Status"] + statuses + 
            [absent_total],
        ]

        # Calculate max width for each column
        max_widths = []
        for col in range(len(header_labels)):
            max_width = max(len(str(header_labels[col])), len(str(rows[1][col])))
            max_widths.append(max_width)

        # Iterate through each row of data and add it to the worksheet
        for i, row_data in enumerate(rows, start=start_row + 1):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.THIN_BORDER
                
                # Style the header column (first columns)
                if col < 7:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    # Adjust column width based on max content width
                    if col <= len(max_widths):
                        ws.column_dimensions[cell.column_letter].width = max_widths[col-1] + 2  # Add padding
                
                # Style the status cells
                if i == start_row + 2:  # Status row
                    style = self.SHIFT_STATUS_STYLES.get(value)
                    if style:
                        cell.style = style

                if i == start_row + 2 and col > len(header_labels) + num_days:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                
                # Style the days row
                if i == start_row + 1 and col > 6: 
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        statuses = []
        absent_total = 0

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                if record.shift_status in ('A', 'MP', 'HD', 'WO'):
                    statuses.append(record.shift_status)
                else:
                    statuses.append("")
            else:
                statuses.append("")

        for status_code in statuses:
            if status_code == 'A':
                absent_total += 1
            elif status_code == 'HD':
                absent_total += 0.5
            elif status_code == 'MP':
                absent_total += 1

        return days, statuses, absent_total
    
class ExportMonthlyAbsentExcel2(View):
    """
    View to generate and export an Excel file containing monthly absent details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        # Add main header
        main_header = f"Monthly Absent Register for the Period of: {first_day_of_month} to {last_day_of_month} | Printed Date: {datetime.now().strftime('%d-%m-%Y')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 6)
        cell = ws.cell(row=1, column=1, value=main_header)
        cell.font = Font(bold=True)

        # Add column headers
        header_labels = ["EMP ID", "EMP Name", "Company", "Location", "Department", "Designation", "Type", "Days"]
        days_header = [str(day) for day in range(1, num_days + 1)]
        total_header = ["Absent Days"]
        full_header = header_labels + days_header + total_header
        ws.append(full_header)

        # Style header row
        for col in range(1, len(full_header) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            cell.border = self.THIN_BORDER

            if col >= 9:  # Style days and total columns
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # Add employee data rows
        row_num = 3  # Start from row 3 after headers
        grand_total = 0

        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                days, statuses, absent_total = self.process_records(records, first_day_of_month, num_days)
                data_row = [
                    details['employee_id'],
                    details['employee_name'],
                    details['company'],
                    details['location'],
                    details['department'],
                    details['designation'],
                    details['job_type'],
                    "Status",
                ] + statuses + [absent_total]
                ws.append(data_row)

                # Apply styling to data row
                for col in range(1, len(data_row) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = self.THIN_BORDER
                    if col <= 8:  # Style employee info columns
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    if col >= 8:  # Style days and total columns
                        cell.font = Font(bold=True)
                        if col < len(data_row):  # Style status cells
                            status = data_row[col - 1]
                            style = self.SHIFT_STATUS_STYLES.get(status)
                            if style:
                                cell.style = style
                        else:  # Style total column
                            cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

                row_num += 1

                # Accumulate grand total seconds
                if absent_total:
                    grand_total += absent_total

        # Add grand total row
        grand_total_row = [""] * (len(full_header) - 2) + ["Grand Total"] + [grand_total]
        ws.append(grand_total_row)

        # Style grand total row
        for col in range(len(full_header) - 1, len(full_header) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color
            cell.border = self.THIN_BORDER

        # Adjust column widths
        self.adjust_column_widths(ws)

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Absent_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        statuses = []
        absent_total = 0

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                if record.shift_status in ('A', 'MP', 'HD', 'WO'):
                    statuses.append(record.shift_status)
                else:
                    statuses.append("")
            else:
                statuses.append("")

        for status_code in statuses:
            if status_code == 'A':
                absent_total += 1
            elif status_code == 'HD':
                absent_total += 0.5
            elif status_code == 'MP':
                absent_total += 1

        return days, statuses, absent_total

    def adjust_column_widths(self, ws):
        """
        Adjust column widths based on the maximum content length in each column,
        skipping merged cells and applying a minimum and maximum width limit.
        """
        max_width = {}
        max_allowed_width = 30  # Set a maximum width limit for columns
        min_allowed_width = 10  # Set a minimum width limit for columns

        for row in ws.iter_rows():
            for cell in row:
                # Skip merged cells as they don't have column_letter
                if isinstance(cell, MergedCell):
                    continue
                column = cell.column_letter
                value_length = len(str(cell.value)) if cell.value is not None else 0
                if column not in max_width or value_length > max_width[column]:
                    max_width[column] = min(value_length, max_allowed_width)  # Apply max width limit

        for col_letter, width in max_width.items():
            # Apply minimum and maximum width, adding padding
            adjusted_width = max(min_allowed_width, width + 2)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    
class ExportMonthlyPresentExcel(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    # Define a border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        row_num = 1
        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                self.add_employee_data(ws, details, records, first_day_of_month, last_day_of_month, num_days, row_num)
                row_num += 4

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Present_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def add_employee_data(self, ws, details, records, first_day, last_day, num_days, start_row):
        """
        Add employee details and attendance data to the worksheet.
        """
        # Add employee header
        header = (
            f"Monthly Late Entry Register for the Period of: {first_day} to {last_day}"
        )
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_days)
        cell = ws.cell(row=start_row, column=1, value=header)
        cell.font = Font(bold=True)

        # Process attendance records
        days, statuses, present_total = self.process_records(records, first_day, num_days)

        # Define header labels
        header_labels = ["EMP ID", "EMP Name", "Department", "Designation", "Type", "Days"]
        row_data = header_labels + days + ["Present Days"]

        # Adding the rows to the sheet
        rows = [
            row_data,
            [f"{details['employee_id']}", f"{details['employee_name']}", f"{details['department']}", f"{details['designation']}", f"{details['job_type']}", "Status"] + statuses + 
            [present_total],
        ]

        # Calculate max width for each column
        max_widths = []
        for col in range(len(header_labels)):
            max_width = max(len(str(header_labels[col])), len(str(rows[1][col])))
            max_widths.append(max_width)

        # Iterate through each row of data and add it to the worksheet
        for i, row_data in enumerate(rows, start=start_row + 1):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.THIN_BORDER
                
                # Style the header column (first columns)
                if col < 7:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    # Adjust column width based on max content width
                    if col <= len(max_widths):
                        ws.column_dimensions[cell.column_letter].width = max_widths[col-1] + 2  # Add padding
                
                # Style the status cells
                if i == start_row + 2:  # Status row
                    style = self.SHIFT_STATUS_STYLES.get(value)
                    if style:
                        cell.style = style

                if i == start_row + 2 and col > len(header_labels) + num_days:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                
                # Style the days row
                if i == start_row + 1 and col > 6: 
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        statuses = []
        present_total = 0

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                if record.shift_status in ('P', 'WW', 'HD', 'WO', 'IH'):
                    statuses.append(record.shift_status)
                else:
                    statuses.append("")
            else:
                statuses.append("")
            
        for status_code in statuses:
            if status_code == 'HD':
                present_total += 0.5
            elif status_code == 'P':
                present_total += 1
            elif status_code == 'WW':
                present_total += 1
            elif status_code == 'IH':
                present_total += 1

        return days, statuses, present_total
    
class ExportMonthlyPresentExcel2(View):
    """
    View to generate and export an Excel file containing monthly duty hours and attendance details for all employees.
    """

    # Predefined styles for shift statuses
    SHIFT_STATUS_STYLES = {
        'P': NamedStyle(name='status-P', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WW': NamedStyle(name='status-WW', fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"), font=Font(color="256029", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'A': NamedStyle(name='status-A', fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"), font=Font(color="C63737", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'HD': NamedStyle(name='status-HD', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'WO': NamedStyle(name='status-WO', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'PH': NamedStyle(name='status-PH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'FH': NamedStyle(name='status-FH', fill=PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid"), font=Font(color="8A5340", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'MP': NamedStyle(name='status-MP', fill=PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"), font=Font(color="D84315", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
        'IH': NamedStyle(name='status-IH', fill=PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), font=Font(color="856404", bold=True), border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), alignment=Alignment(horizontal='center')),
    }

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def get(self, request, *args, **kwargs):
        """
        Handle GET request to generate the Excel file.
        """
        month = request.GET.get('month')
        year = request.GET.get('year')

        if not (month and year):
            return HttpResponse({'error': 'Month and year are required parameters.'}, status=400)

        try:
            month, year = int(month), int(year)
        except ValueError:
            return HttpResponse({'error': 'Month and year must be valid integers.'}, status=400)

        # Fetch attendance and employee data
        attendance_data, employee_data = self.fetch_data(month, year)

        # Initialize the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{month}_{year}"

        first_day_of_month = date(year, month, 1)
        last_day_of_month = self.get_last_day_of_month(first_day_of_month)
        num_days = last_day_of_month.day

        # Add the headers for the file
        main_header = f"Monthly Present Register for the Period: {first_day_of_month} to {last_day_of_month} | Printed: {datetime.now().strftime('%d-%m-%Y')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 6)
        cell = ws.cell(row=1, column=1, value=main_header)
        cell.font = Font(bold=True)

        # Add column headers
        header_labels = ["EMP ID", "EMP Name", "Company", "Location", "Department", "Designation", "Type", "Days"]
        days_header = [str(day) for day in range(1, num_days + 1)]
        total_header = ["Present Days"]
        full_header = header_labels + days_header + total_header
        ws.append(full_header)

        # Style header row
        for col in range(1, len(full_header) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            cell.border = self.THIN_BORDER

            if col >= 9:  # Style days and total columns
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # Add employee data rows
        row_num = 3  # Start from row 3 after headers
        grand_total = 0

        for employee_id, details in employee_data.items():
            records = attendance_data.get(employee_id, [])
            if records:
                days, statuses, present_total = self.process_records(records, first_day_of_month, num_days)
                data_row = [
                    details['employee_id'],
                    details['employee_name'],
                    details['company'],
                    details['location'],
                    details['department'],
                    details['designation'],
                    details['job_type'],
                    "Status",
                ] + statuses + [present_total]
                ws.append(data_row)

                # Apply styling to data row
                for col in range(1, len(data_row) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = self.THIN_BORDER
                    if col <= 8:  # Style employee info columns
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

                    if col >= 8:  # Style days and total columns
                        if col < len(data_row):  # Style status cells
                            status = data_row[col - 1]
                            style = self.SHIFT_STATUS_STYLES.get(status)
                            if style:
                                cell.style = style
                        else:  # Style total column
                            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    
                    if col > len(header_labels) + num_days:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

                row_num += 1

                # Accumulate grand total seconds
                if present_total:
                    grand_total += present_total

        # Add grand total row
        grand_total_row = [""] * (len(full_header) - 2) + ["Grand Total"] + [grand_total]
        ws.append(grand_total_row)

        # Style grand total row
        for col in range(len(full_header) - 1, len(full_header) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color
            cell.border = self.THIN_BORDER

        # Adjust column widths
        self.adjust_column_widths(ws)

        # Generate response
        month_name = first_day_of_month.strftime('%B')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Employee_Present_{month_name}_{year}.xlsx'
        wb.save(response)
        return response

    def fetch_data(self, month, year):
        """
        Fetch attendance and employee data to minimize database queries.
        """
        # Fetch attendance records
        attendance_records = Attendance.objects.filter(logdate__year=year, logdate__month=month).select_related('employeeid')
        attendance_data = defaultdict(list)
        for record in attendance_records:
            attendance_data[record.employeeid_id].append(record)

        # Fetch employee data
        employees = Employee.objects.all().select_related('department', 'designation', 'location', 'company')
        employee_data = {
            emp.id: {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'company': emp.company.name,
                'location': emp.location.name,
                'department': emp.department.name,
                'designation': emp.designation.name,
                'job_type': emp.job_type,
                'location': emp.location.name,
                'company': emp.company.name,
            }
            for emp in employees
        }

        return attendance_data, employee_data

    def get_last_day_of_month(self, first_day):
        """
        Get the last day of the month.
        """
        next_month = first_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    def add_employee_data(self, ws, details, records, first_day, last_day, num_days, start_row):
        """
        Add employee details and attendance data to the worksheet.
        """
        # Add employee header
        header = (
            f"Monthly Late Entry Register for the Period of: {first_day} to {last_day}"
        )
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_days)
        cell = ws.cell(row=start_row, column=1, value=header)
        cell.font = Font(bold=True)

        # Process attendance records
        days, statuses, present_total = self.process_records(records, first_day, num_days)

        # Define header labels
        header_labels = ["EMP ID", "EMP Name", "Company", "Location", "Department", "Designation", "Type", "Days"]
        row_data = header_labels + days + ["Present Days"]

        # Adding the rows to the sheet
        rows = [
            row_data,
            [f"{details['employee_id']}", f"{details['employee_name']}", f"{details['department']}", f"{details['designation']}", f"{details['job_type']}", "Status"] + statuses + 
            [present_total],
        ]

        # Calculate max width for each column
        max_widths = []
        for col in range(len(header_labels)):
            max_width = max(len(str(header_labels[col])), len(str(rows[1][col])))
            max_widths.append(max_width)

        # Iterate through each row of data and add it to the worksheet
        for i, row_data in enumerate(rows, start=start_row + 1):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.THIN_BORDER
                
                # Style the header column (first columns)
                if col < 7:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    # Adjust column width based on max content width
                    if col <= len(max_widths):
                        ws.column_dimensions[cell.column_letter].width = max_widths[col-1] + 2  # Add padding
                
                # Style the status cells
                if i == start_row + 2:  # Status row
                    style = self.SHIFT_STATUS_STYLES.get(value)
                    if style:
                        cell.style = style

                if i == start_row + 2 and col > len(header_labels) + num_days:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                
                # Style the days row
                if i == start_row + 1 and col > 6: 
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    def process_records(self, records, first_day, num_days):
        """
        Process attendance records for each day of the month.
        """
        days = [str(day) for day in range(1, num_days + 1)]
        statuses = []
        present_total = 0

        for day in range(1, num_days + 1):
            log_date = first_day.replace(day=day)
            record = next((rec for rec in records if rec.logdate == log_date), None)
            if record:
                if record.shift_status in ('P', 'WW', 'HD', 'WO', 'IH'):
                    statuses.append(record.shift_status)
                else:
                    statuses.append("")
            else:
                statuses.append("")
            
        for status_code in statuses:
            if status_code == 'HD':
                present_total += 0.5
            elif status_code == 'P':
                present_total += 1
            elif status_code == 'WW':
                present_total += 1
            elif status_code == 'IH':
                present_total += 1

        return days, statuses, present_total

    def adjust_column_widths(self, ws):
        """
        Adjust column widths based on the maximum content length in each column,
        skipping merged cells and applying a minimum and maximum width limit.
        """
        max_width = {}
        max_allowed_width = 30  # Set a maximum width limit for columns
        min_allowed_width = 10  # Set a minimum width limit for columns

        for row in ws.iter_rows():
            for cell in row:
                # Skip merged cells as they don't have column_letter
                if isinstance(cell, MergedCell):
                    continue
                column = cell.column_letter
                value_length = len(str(cell.value)) if cell.value is not None else 0
                if column not in max_width or value_length > max_width[column]:
                    max_width[column] = min(value_length, max_allowed_width)  # Apply max width limit

        for col_letter, width in max_width.items():
            # Apply minimum and maximum width, adding padding
            adjusted_width = max(min_allowed_width, width + 2)
            ws.column_dimensions[col_letter].width = adjusted_width


class ProcessLogView(APIView):
    def post(self, request, *args, **kwargs):
        serializer = serializers.LogsSerializer(data=request.data)
        if serializer.is_valid():
            log_datetime = serializer.validated_data['log_datetime']
            employeeid = serializer.validated_data['employeeid']
            direction = serializer.validated_data['direction']

            # Create a Logs instance
            log = Logs(log_datetime=log_datetime, employeeid=employeeid, direction=direction)

            # Process the log using AttendanceProcessor
            processor = attendance5.AttendanceProcessor()
            success = processor.process_single_log(log, is_manual=True)

            if success:
                return Response({"status": "success", "message": "Log processed successfully."}, status=status.HTTP_200_OK)
            else:
                return Response({"status": "error", "message": "Failed to process log."}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class MonthlyAttendanceView(APIView):
    """
    API to fetch attendance data for a given year and month.
    Uses pagination for employee records.
    """
    class EmployeePagination(PageNumberPagination):
        page_size = 10  # Customize the page size
        page_size_query_param = 'page_size'
        max_page_size = 50

    def get(self, request):
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        search_query = request.query_params.get('search', '')

        if not (year and month):
            return Response({'error': 'Year and month are required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            year, month = int(year), int(month)
            if month < 1 or month > 12:
                raise ValueError("Invalid month")
        except ValueError:
            return Response({'error': 'Invalid year or month'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get the current date
        today = datetime.today()

        # Get the number of days in the given month
        days_in_month = monthrange(year, month)[1]

        # If the requested month is the current month, limit the days to today’s date
        if year == today.year and month == today.month:
            days_in_month = today.day  # Limit to current date

        # Generate dates list
        dates_list = [
            {
                'date': f"{day:02d}",
                'day': datetime(year, month, day).strftime('%A')
            }
            for day in range(1, days_in_month + 1)
        ]

        # Get attendance records for the given month
        start_date = make_aware(datetime(year, month, 1))
        end_date = make_aware(datetime(year, month, days_in_month, 23, 59, 59))

        attendance_records = Attendance.objects.filter(logdate__range=(start_date, end_date))

        # Apply search filter for employee_id and employee_name
        if search_query:
            attendance_records = attendance_records.filter(
                Q(employeeid__employee_id__icontains=search_query) 
                # Q(employeeid__employee_name__icontains=search_query)
            )

        # Group attendance by employee
        employee_data = {}
        for record in attendance_records:
            emp_id = record.employeeid.employee_id
            emp_name = record.employeeid.employee_name

            if emp_id not in employee_data:
                employee_data[emp_id] = {
                    'employee_id': emp_id,
                    'employee_name': emp_name,
                    'attendance': {}
                }

            date_str = record.logdate.strftime('%d')
            employee_data[emp_id]['attendance'][date_str] = {
                'id': record.id,
                'time_in': record.first_logtime.strftime('%H:%M') if record.first_logtime else None,
                'time_out': record.last_logtime.strftime('%H:%M') if record.last_logtime else None,
                'direction': record.direction if record.direction else None
            }

        # Paginate employee records
        paginator = self.EmployeePagination()
        paginated_employees = paginator.paginate_queryset(list(employee_data.values()), request)

        response_data = {
            'dates': dates_list,
            'employees': paginated_employees
        }

        return paginator.get_paginated_response(response_data)

# class MonthlyAttendanceView(APIView):
#     """
#     API to fetch attendance data for a given year and month.
#     Optimized for faster execution by reducing database queries.
#     """
#     class EmployeePagination(PageNumberPagination):
#         page_size = 10  # Customize the page size
#         page_size_query_param = 'page_size'
#         max_page_size = 50

#     def get(self, request):
#         year = request.query_params.get('year')
#         month = request.query_params.get('month')
#         search_query = request.query_params.get('search', '').strip()

#         if not (year and month):
#             return Response({'error': 'Year and month are required'}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             year, month = int(year), int(month)
#             if month < 1 or month > 12:
#                 raise ValueError("Invalid month")
#         except ValueError:
#             return Response({'error': 'Invalid year or month'}, status=status.HTTP_400_BAD_REQUEST)

#         today = datetime.today()
#         days_in_month = monthrange(year, month)[1]

#         # If the requested month is the current month, limit the days to today's date
#         if year == today.year and month == today.month:
#             days_in_month = today.day  

#         # Generate dates list
#         dates_list = [
#             {'date': f"{day:02d}", 'day': datetime(year, month, day).strftime('%A')}
#             # {'date': f"{day:02d}-{month:02d}-{year:02d}", 'day': datetime(year, month, day).strftime('%A')}
#             for day in range(1, days_in_month + 1)
#         ]

#         # Optimize DB query: Only fetch required fields and related employee data in one query
#         start_date = make_aware(datetime(year, month, 1))
#         end_date = make_aware(datetime(year, month, days_in_month, 23, 59, 59))

#         attendance_query = Attendance.objects.filter(logdate__range=(start_date, end_date)).select_related(
#             'employeeid'
#         ).only(
#             'id', 'logdate', 'first_logtime', 'last_logtime', 'direction',
#             'employeeid__employee_id', 'employeeid__employee_name'
#         )

#         # Apply search filter efficiently
#         if search_query:
#             attendance_query = attendance_query.filter(
#                 Q(employeeid__employee_id__icontains=search_query) |
#                 Q(employeeid__employee_name__icontains=search_query)
#             )

#         # Convert QuerySet to list of dictionaries (Bulk fetch)
#         attendance_list = list(attendance_query.values(
#             'id', 'logdate', 'first_logtime', 'last_logtime', 'direction',
#             'employeeid__employee_id', 'employeeid__employee_name'
#         ))

#         # Grouping employees efficiently
#         employee_data = {}
#         for record in attendance_list:
#             emp_id = str(record['employeeid__employee_id'])
#             emp_name = record['employeeid__employee_name']

#             if emp_id not in employee_data:
#                 employee_data[emp_id] = {
#                     'employee_id': emp_id,
#                     'employee_name': emp_name,
#                     'attendance': {}
#                 }

#             date_str = record['logdate'].strftime('%d')
#             employee_data[emp_id]['attendance'][date_str] = {
#                 'id': record['id'],
#                 'time_in': record['first_logtime'].strftime('%H:%M') if record['first_logtime'] else None,
#                 'time_out': record['last_logtime'].strftime('%H:%M') if record['last_logtime'] else None,
#                 'direction': record['direction'] if record['direction'] else None
#             }

#         # Paginate employee records
#         paginator = self.EmployeePagination()
#         paginated_employees = paginator.paginate_queryset(list(employee_data.values()), request)

#         response_data = {
#             'dates': dates_list,
#             'employees': paginated_employees
#         }

#         return paginator.get_paginated_response(response_data)
    
# class UpdateAttendanceView(APIView):
#     def patch(self, request, attendance_id):
#         from resource.attendance5 import AttendanceProcessor
#         print("UpdateAttendanceView PATCH method called!")  # Debugging statement

#         processor = AttendanceProcessor()  # Create instance once
#         data = request.data
#         time_in = data.get("time_in")
#         time_out = data.get("time_out")

#         if not time_in and not time_out:
#             return Response({"error": "Either time_in or time_out must be provided."}, status=status.HTTP_400_BAD_REQUEST)

#         attendance = get_object_or_404(Attendance, id=attendance_id)
#         employeeid = attendance.employeeid.employee_id
#         log_date = attendance.logdate
#         log_entries = []
#         logs_to_reprocess = []

#         # Convert times to HH:MM format
#         formatted_time_in = None
#         formatted_time_out = None

#         if time_in:
#             time_in = parser.parse(time_in).time()
#             formatted_time_in = time_in.strftime("%H:%M")
#             log_datetime_in = make_aware(datetime.combine(log_date, time_in))

#             # Log "In Device"
#             log_entry_in = ManualLogs(employeeid=employeeid, log_datetime=log_datetime_in, direction="In Device")
#             log_entries.append(log_entry_in)
#             logs_to_reprocess.append(log_entry_in)

#             # If last_logtime is NOT null, reprocess "Out Device" log
#             if attendance.last_logtime:
#                 log_datetime_out = make_aware(datetime.combine(log_date, attendance.last_logtime))
#                 log_entry_out = ManualLogs(employeeid=employeeid, log_datetime=log_datetime_out, direction="Out Device")

#                 log_entries.append(log_entry_out)
#                 logs_to_reprocess.append(log_entry_out)  # Ensure reprocessing

#         if time_out:
#             time_out = parser.parse(time_out).time()
#             formatted_time_out = time_out.strftime("%H:%M")
#             log_datetime_out = make_aware(datetime.combine(log_date, time_out))

#             # Log "Out Device"
#             log_entry_out = ManualLogs(employeeid=employeeid, log_datetime=log_datetime_out, direction="Out Device")
#             log_entries.append(log_entry_out)
#             logs_to_reprocess.append(log_entry_out)  # Ensure reprocessing

#         # Save logs
#         ManualLogs.objects.bulk_create(log_entries)

#         # Ensure **reprocessing** of all relevant logs
#         for log in logs_to_reprocess:
#             print(f"Reprocessing log for {log.employeeid} at {log.log_datetime}")
#             processor.process_single_log(log, is_manual=True)

#         return Response({
#             "message": "Attendance record updated successfully.",
#             "time_in": formatted_time_in if time_in else None,
#             "time_out": formatted_time_out if time_out else None
#         }, status=status.HTTP_200_OK)

class UpdateAttendanceView(APIView):
    def patch(self, request, attendance_id):
        from resource.attendance7 import AttendanceProcessor
        print("UpdateAttendanceView PATCH method called!")

        processor = AttendanceProcessor()
        data = request.data
        time_in = data.get("time_in")
        time_out = data.get("time_out")

        if not time_in and not time_out:
            return Response({"error": "Either time_in or time_out must be provided."}, status=status.HTTP_400_BAD_REQUEST)

        attendance = get_object_or_404(Attendance, id=attendance_id)
        employeeid = attendance.employeeid.employee_id
        log_date = attendance.logdate
        log_entries = []
        logs_to_reprocess = []

        formatted_time_in = None
        formatted_time_out = None

        if time_in and attendance.last_logtime is None:
            time_in_time = parser.parse(time_in).time()  # Parse time only once
            formatted_time_in = time_in_time.strftime("%H:%M")
            log_datetime_in = make_aware(datetime.combine(log_date, time_in_time))

            # Log "In Device"
            log_entry_in = ManualLogs(employeeid=employeeid, log_datetime=log_datetime_in, direction="In Device")
            log_entries.append(log_entry_in)
            logs_to_reprocess.append(log_entry_in)
            print(f"DEBUG: Created IN log: {log_entry_in.log_datetime}") # Debug print

            attendance.first_logtime = time_in_time  # Directly update first_logtime
            attendance.direction = "Manual"  # Set direction to "Manual" for manual logs
            attendance.save() # Save immediately after updating first_logtime
            print(f"DEBUG: Updated Attendance first_logtime to: {attendance.first_logtime}") # Debug print


        if time_out:
            time_out_time = parser.parse(time_out).time() # Parse time only once
            formatted_time_out = time_out_time.strftime("%H:%M")
            log_datetime_out = make_aware(datetime.combine(log_date, time_out_time))

            # Log "Out Device"
            log_entry_out = ManualLogs(employeeid=employeeid, log_datetime=log_datetime_out, direction="Out Device")
            log_entries.append(log_entry_out)
            logs_to_reprocess.append(log_entry_out) # Ensure reprocessing
            print(f"DEBUG: Created OUT log: {log_entry_out.log_datetime}") # Debug print

            attendance.last_logtime = time_out_time # Directly update last_logtime
            attendance.direction = "Manual"
            attendance.save() # Save immediately after updating last_logtime
            print(f"DEBUG: Updated Attendance last_logtime to: {attendance.last_logtime}") # Debug print


        elif time_in and attendance.last_logtime is not None and attendance.first_logtime is None:
            # EDGE CASE: first_logtime is null, last_logtime exists, adding time_in
            print("DEBUG: Edge case - reprocessing OUT log based on existing last_logtime") # Debug print

            # No need to create a NEW ManualLogs for OUT, reuse existing last_logtime
            log_datetime_out_exist = make_aware(datetime.combine(log_date, attendance.last_logtime))
            # Create a ManualLogs entry for the *existing* last_logtime to trigger reprocessing
            log_entry_out_exist = ManualLogs(employeeid=employeeid, log_datetime=log_datetime_out_exist, direction="Out Device")
            log_entries.append(log_entry_out_exist) # Still create for log consistency if needed, but may not be strictly necessary
            logs_to_reprocess.append(log_entry_out_exist) # Ensure reprocessing of OUT


            time_in_time = parser.parse(time_in).time()
            formatted_time_in = time_in_time.strftime("%H:%M")
            log_datetime_in = make_aware(datetime.combine(log_date, time_in_time))

            # Log "In Device"
            log_entry_in = ManualLogs(employeeid=employeeid, log_datetime=log_datetime_in, direction="In Device")
            log_entries.append(log_entry_in)
            logs_to_reprocess.append(log_entry_in)
            print(f"DEBUG: Created IN log (edge case): {log_entry_in.log_datetime}") # Debug print

            attendance.first_logtime = time_in_time # Directly update first_logtime
            attendance.direction = "Manual"
            attendance.save() # Save immediately
            print(f"DEBUG: Updated Attendance first_logtime (edge case) to: {attendance.first_logtime}") # Debug print


        # Save logs (bulk create is efficient, even if sometimes list is short)
        ManualLogs.objects.bulk_create(log_entries)
        print(f"DEBUG: Saved ManualLogs: {[log.direction for log in log_entries]}") # Debug print


        # Reprocess logs - IMPORTANT: Process OUT *after* IN in the edge case!
        if time_in and attendance.last_logtime and not attendance.first_logtime:
            print("DEBUG: Reprocessing IN log (edge case) first...") # Debug print
            processor.process_single_log(logs_to_reprocess[-1], is_manual=True) # Process IN log (which is now the last one in list)
            print("DEBUG: Reprocessing OUT log (edge case) second...") # Debug print
            processor.process_single_log(logs_to_reprocess[0], is_manual=True) # Process OUT log (which was added first in this case)

        else:
            print("DEBUG: Reprocessing logs in order...") # Debug print
            for log in logs_to_reprocess:
                print(f"DEBUG: Reprocessing log: {log.direction} at {log.log_datetime}") # Debug print
                processor.process_single_log(log, is_manual=True)


        return Response({
            "message": "Attendance record updated successfully.",
            "time_in": formatted_time_in if time_in else None,
            "time_out": formatted_time_out if time_out else None
        }, status=status.HTTP_200_OK)
    
class HolidayListCreate(generics.ListCreateAPIView):
    """
    API view for listing and creating logs with improved filtering.
    """
    serializer_class = serializers.HolidayListSerializer
    pagination_class = DefaultPagination
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    filter_fields = '__all__'
    search_fields = ['holiday_date', 'holiday_name', 'holiday_type']
    ordering_fields = ['holiday_date', 'holiday_name']
    ordering_fields = '__all__'
    ordering = ['-holiday_date']

    def get_queryset(self):
        """
        Get the queryset for listing logs with all filters.
        """
        queryset = HolidayList.objects.all()

        # Get query parameters
        search_query = self.request.GET.get('search')
        
        # Apply search filter (case-insensitive)
        if search_query:
            queryset = queryset.filter(
                Q(holiday_date__icontains=search_query) |
                Q(holiday_name__icontains=search_query) |
                Q(holiday_type__icontains=search_query)
            )

        return queryset.order_by('-holiday_date')

    def get(self, request, *args, **kwargs):
        """
        Get the list of Logs records with filters.
        """
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
            
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)        

class HolidayRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    """
    API view for retrieving, updating, and deleting a log.
    """
    queryset = HolidayList.objects.all()
    serializer_class = serializers.HolidayListSerializer
    lookup_url_kwarg = "id"